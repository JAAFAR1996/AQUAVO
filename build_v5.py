import openpyxl, math, io, sys
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

USD = 1520
SHIP1 = 600 * USD  # عادي 912,000
SHIP2 = 500 * USD  # حساس 760,000
UGC_TOTAL = 70 * USD  # 1000 بطاقة = 106,400
CARTON = {'S':500,'M':800,'L':1000}
# هوامش متفاوتة حسب الفئة (استراتيجية اختراق السوق)
# استهلاكي سريع الدوران: 30-35% (منتجات المقارنة)
# علاجي/فيتامينات: 60-80% (قيمة عالية عند المربي، تكلفة منخفضة)
# معدات (أحواض/فلاتر): 40-45% (منتجات متوسطة)
MARGINS = {
    'غذاء':      0.32,   # طعام سمك — سريع الدوران
    'سخان':      0.33,   # سخانات — منتج مقارنة
    'فلتر':      0.35,   # فلاتر ومضخات
    'حوض':       0.42,   # أحواض زجاجية — استثمار
    'علاج':      0.70,   # أدوية (تكلفة قليلة، قيمة عالية)
    'فيتامين':   0.65,   # أملاح ومعادن
    'اختبار':    0.60,   # شرائط فحص ومحاليل
    'كيماوي':    0.55,   # مثبتات ومزيلات
    'تربة':      0.45,   # تربة أحواض
    'إكسسوار':   0.40,   # إضاءة، فرش، خراطيم
    'هدية':      0.00,   # منتجات مجانية
}
ENDS = [50,100,150,200,250,300,350,400,450,550,600,650,700,750,800,850,900,950]

def get_cat(en, cn):
    s = (en+cn).lower()
    # هدايا مجانية
    if any(k in s for k in ['free','gift','广角灯','磁力刷','牛筋管','适配器','adapter']): return 'هدية'
    # غذاء
    if any(k in s for k in ['feed','food','brine','worm','shrimp','饲料','粮','卵','冻干','鱼粮']): return 'غذاء'
    # سخانات
    if any(k in s for k in ['heat','ysh','quartz','加热','武士']): return 'سخان'
    # فلاتر ومضخات
    if any(k in s for k in ['filter','pump','oil film','换水','过滤','油膜','滤材','培养环','培菌']): return 'فلتر'
    # أحواض
    if any(k in s for k in ['tank','mm thick','glass','缸','cm 6mm','cm 8mm','cm 5mm']): return 'حوض'
    # علاج طبي
    if any(k in s for k in ['spot','methylene','bactericid','白点','黄粉','亚甲基蓝']): return 'علاج'
    # فيتامينات وأملاح
    if any(k in s for k in ['mineral','salt','矿物盐','多维']): return 'فيتامين'
    # اختبارات
    if any(k in s for k in ['test','tester','试纸','测试']): return 'اختبار'
    # كيماويات
    if any(k in s for k in ['probio','nitrif','stabilizer','algae','chlorine','ammonia',
                              'bacteria','descal','除','净','稳定','硝化','益生','藻']): return 'كيماوي'
    # تربة
    if any(k in s for k in ['grass mud','水草泥']): return 'تربة'
    # إكسسوار
    if any(k in s for k in ['light','brush','tube','气泵','气盘','隔离','孵化','培养箱',
                              'thermometer','温度计','气囊']): return 'إكسسوار'
    return 'إكسسوار'

def pprice(cost, margin):
    mp = cost / (1 - margin)
    tail = int(cost) % 1000
    pref = min(ENDS, key=lambda e: abs(e - tail))
    bk = int(math.ceil(mp / 1000))
    c = bk * 1000 - 1000 + pref
    if c < mp: c = bk * 1000 + pref
    if c < mp: c += 1000
    return int(c)

# 1. قراءة أسعار الشراء من 1.5
src15 = openpyxl.load_workbook(
    r'c:\Users\jaafa\Desktop\upload\FishWebClean\客户伊拉克-Jaafar-1.5 (1).xlsx', data_only=True)
ws15 = src15.active
prices = {}  # code -> unit_price_usd
for row in ws15.iter_rows(values_only=True):
    code = str(row[2]).strip() if row[2] else ''
    if not code or code in ('','Business\nCode'): continue
    try:
        up = float(str(row[6]).replace(',',''))
        prices[code] = up
    except: pass

# 2. قراءة 1.9 عادي (Order 1)
wb1 = openpyxl.load_workbook(
    r'C:\Users\jaafa\Downloads\ros\伊拉克-Jaafar-1.9 - packing list（普货）.xlsx', data_only=True)
items1 = []
tnw1 = 0
for row in wb1.active.iter_rows(values_only=True):
    try: num = int(row[0])
    except: continue
    code = str(row[2]).strip()
    cn = str(row[3])[:25] if row[3] else ''
    en = str(row[4])[:40] if row[4] else ''
    qty = int(row[5]) if row[5] else 0
    unw = float(row[9]) if row[9] else 0
    tnw1 += qty * unw
    items1.append({'code':code,'cn':cn,'en':en,'qty':qty,'unw':unw,'tnw':qty*unw})

# 3. قراءة 1.9 حساس (Order 2)
wb2 = openpyxl.load_workbook(
    r'C:\Users\jaafa\Downloads\ros\伊拉克-Jaafar-1.9 - packing list（敏货）.xlsx', data_only=True)
items2 = []
tnw2 = 0
for row in wb2.active.iter_rows(values_only=True):
    try: num = int(row[0])
    except: continue
    code = str(row[2]).strip()
    cn = str(row[3])[:25] if row[3] else ''
    en = str(row[4])[:40] if row[4] else ''
    qty = int(row[5]) if row[5] else 0
    unw = float(row[9]) if row[9] else 0
    tnw2 += qty * unw
    items2.append({'code':code,'cn':cn,'en':en,'qty':qty,'unw':unw,'tnw':qty*unw})

# === التوزيع النسبي للقيمة (Value-Based Allocation) ===
# الأحواض = صناديق شحن مجانية للمنتجات الصغيرة (打木箱，内部塞其它产品)
# لذلك نوزع الشحن بنسبة قيمة كل منتج من إجمالي الفاتورة
# وليس بالوزن (الذي يظلم الأحواض الثقيلة)

# حساب إجمالي قيمة كل طلب
tval1 = 0
for it in items1:
    up = prices.get(it['code'], 0)
    tval1 += up * it['qty']
tval2 = 0
for it in items2:
    up = prices.get(it['code'], 0)
    tval2 += up * it['qty']

total_units = sum(i['qty'] for i in items1+items2)
ugc_per = round(UGC_TOTAL / total_units)

src_upd = openpyxl.load_workbook(
    r'c:\Users\jaafa\Desktop\upload\FishWebClean\Yee_Products_2026_UPDATED.xlsx', data_only=True)
box_map = {}
for row in src_upd.active.iter_rows(values_only=True):
    if not row[0]: continue
    cn2 = str(row[1])[:15] if row[1] else ''
    box_map[cn2] = str(row[5]) if row[5] else 'S'

products = []
for order, items, ship_total, tval in [('عادي',items1,SHIP1,tval1),('حساس',items2,SHIP2,tval2)]:
    for it in items:
        up = prices.get(it['code'], 0)
        box = 'S'
        for k,v in box_map.items():
            if k and k in it['cn']: box = v; break
        carton = CARTON.get(box, 500)
        # شحن بنسبة القيمة: (قيمة المنتج / إجمالي الفاتورة) × إجمالي الشحن
        product_val = up * it['qty']
        val_ratio = product_val / tval if tval > 0 else 0
        ship_product = ship_total * val_ratio  # شحن كل الكمية
        ship_u = round(ship_product / it['qty']) if it['qty'] > 0 else 0
        cat = get_cat(it['en'], it['cn'])
        margin = MARGINS.get(cat, 0.40)
        cost = round(up * USD + carton + ship_u + ugc_per)
        if cat == 'هدية':
            sell = 0; mg = 0; profit = 0
        else:
            sell = pprice(cost, margin)
            mg = round((sell-cost)/sell*100, 1) if sell>0 else 0
            profit = sell - cost
        products.append({
            'en':it['en'],'cn':it['cn'],'order':order,'cat':cat,
            'qty':it['qty'],'buy':up,'buy_iqd':round(up*USD),
            'box':box,'carton':carton,'ship':ship_u,'ugc':ugc_per,
            'cost':cost,'sell':sell,'margin':mg,'target_mg':round(margin*100),
            'profit':profit,'total_p':profit*it['qty'],
            'pts':sell//1000,'tail':sell%1000,
            'unw':it['unw'],
        })

trev = sum(p['sell']*p['qty'] for p in products)
tcost = sum(p['cost']*p['qty'] for p in products)
tprof = trev - tcost
amg = tprof/trev*100

# Excel
def hf(c): return PatternFill('solid',fgColor=c)
def ft(b=False,c='000000',s=9,i=False): return Font(bold=b,color=c,size=s,italic=i,name='Calibri')
def al(h='center',v='center',w=False): return Alignment(horizontal=h,vertical=v,wrap_text=w,readingOrder=2)
T=Side(style='thin',color='CCCCCC')
def bd(): return Border(top=T,bottom=T,left=T,right=T)
NV='1A1A2E';RD='E63946';GD='FFD700';TL='A8DADC';MT='D8F3DC';AM='FFF3CD';RS='FFD6D6';LG='F5F5F5'

wb=openpyxl.Workbook()
ws=wb.active; ws.title='التسعير'; ws.sheet_view.rightToLeft=True; ws.freeze_panes='A4'

ws.merge_cells('A1:U1')
ws['A1'].value='AQUAVO — تسعير نهائي 2026 | هوامش متفاوتة | شحن بنسبة القيمة (Value-Based) | 1$=1,520د'
ws['A1'].fill=hf(NV); ws['A1'].font=ft(True,TL,11); ws['A1'].alignment=al()

ws.merge_cells('A2:U2')
ws['A2'].value=f'غذاء:32% | سخان:33% | فلتر:35% | حوض:42% | تربة:45% | كيماوي:55% | اختبار:60% | فيتامين:65% | علاج:70% | شحن عادي:{SHIP1:,}د على {tval1:.0f}$ | حساس:{SHIP2:,}د على {tval2:.0f}$'
ws['A2'].fill=hf('16213E'); ws['A2'].font=ft(False,TL,7,True); ws['A2'].alignment=al()

H=[('رقم',3),('المنتج',34),('الفئة',8),('طلب',5),('كمية',4),('شراء$',6),('شراء د',8),
   ('وزن',5),('كرتون',6),('شحن/ق',7),('UGC',4),('التكلفة',9),
   ('سعر بيع',9),('هدف%',5),('هامش%',6),('ربح/ق',8),('ربح كلي',9),
   ('خصم5%',6),('خصم10%',6),('نقاط',5),('ذيل',5)]
for ci,(h,w) in enumerate(H,1):
    c=ws.cell(3,ci,h); c.fill=hf(RD); c.font=ft(True,'FFFFFF',7); c.alignment=al(w=True); c.border=bd()
    ws.column_dimensions[get_column_letter(ci)].width=w

for i,p in enumerate(products):
    r=i+4; bg=LG if i%2==0 else 'FFFFFF'
    mf=MT if p['margin']>=p['target_mg'] else AM if p['margin']>=p['target_mg']-10 else RS
    s=p['sell']
    d5mg=round((s*.95-p['cost'])/(s*.95)*100,1) if s>0 else 0
    d10mg=round((s*.90-p['cost'])/(s*.90)*100,1) if s>0 else 0
    vals=[i+1,p['en'],p['cat'],p['order'],p['qty'],round(p['buy'],2),p['buy_iqd'],
          p['unw'],p['carton'],p['ship'],p['ugc'],p['cost'],
          s,p['target_mg'],p['margin'],p['profit'],p['total_p'],
          d5mg,d10mg,int(p['pts']),int(p['tail'])]
    for ci,v in enumerate(vals,1):
        cell=ws.cell(r,ci,v); cell.border=bd(); cell.font=ft(s=7); cell.alignment=al(h='right',w=(ci==2))
        if ci==13: cell.fill=hf('E8F4FD'); cell.font=ft(True,'003366',9); cell.number_format='#,##0'
        elif ci==3: cell.fill=hf('FFF0E0'); cell.font=ft(True,s=7)  # فئة
        elif ci==14: cell.fill=hf('E0E0FF')  # هدف
        elif ci==15: cell.fill=hf(mf); cell.font=ft(True,s=8)
        elif ci==10: cell.fill=hf('FFF0F0')
        elif ci in(18,19): cell.fill=hf(MT) if isinstance(v,(int,float)) and v>=20 else hf(AM) if isinstance(v,(int,float)) and v>=10 else hf(RS)
        else: cell.fill=hf(bg)

tr=len(products)+4
ws.merge_cells(f'A{tr}:L{tr}')
ws.cell(tr,1,'الإجماليات').fill=hf(NV); ws.cell(tr,1).font=ft(True,GD,9); ws.cell(tr,1).alignment=al()
for ci,v in [(13,int(trev)),(15,round(amg,1)),(17,int(tprof))]:
    c=ws.cell(tr,ci,v); c.fill=hf(NV); c.font=ft(True,GD,9); c.alignment=al(); c.border=bd()

# ملخص
ws2=wb.create_sheet('الملخص'); ws2.sheet_view.rightToLeft=True
ws2['A1'].value='الملخص المالي'; ws2['A1'].fill=hf(NV); ws2['A1'].font=ft(True,TL,12)
ws2.merge_cells('A1:C1')
data=[('','دينار','دولار'),
('شحن عادي (0.92CBM)',f'{SHIP1:,}','$600'),('شحن حساس (0.37CBM)',f'{SHIP2:,}','$500'),
('UGC 1000 بطاقة',f'{UGC_TOTAL:,}','$70'),
('','',''),('إجمالي الإيرادات',f'{int(trev):,}',f'${int(trev/USD):,}'),
('صافي الربح',f'{int(tprof):,}',f'${int(tprof/USD):,}'),
('هامش الربح',f'{amg:.1f}%',''),
('ربح بعد خصم 5%',f'{int(trev*.95-tcost):,}',''),
('ربح بعد خصم 10%',f'{int(trev*.90-tcost):,}','')]
for ri,row in enumerate(data,2):
    for ci,v in enumerate(row,1):
        c=ws2.cell(ri,ci,v); c.border=bd(); c.font=ft(s=9); c.alignment=al(h='right')
ws2.column_dimensions['A'].width=25; ws2.column_dimensions['B'].width=18; ws2.column_dimensions['C'].width=12

out=r'c:\Users\jaafa\Desktop\upload\FishWebClean\AQUAVO_v6_FINAL.xlsx'
wb.save(out)
o=io.StringIO()
o.write(f'SAVED: {out}\n')
o.write(f'Products: {len(products)} | Rev: {int(trev):,} | Profit: {int(tprof):,} | Margin: {amg:.1f}%\n\n')
# Show by category
from collections import defaultdict
by_cat = defaultdict(list)
for p in products: by_cat[p['cat']].append(p)
for cat in sorted(by_cat.keys()):
    items = by_cat[cat]
    o.write(f'--- {cat} (هدف {items[0]["target_mg"]}%) ---\n')
    for p in items[:4]:
        o.write(f'  {p["en"][:30]:<30} cost:{p["cost"]:>6} sell:{p["sell"]:>6} mg:{p["margin"]}%\n')
    if len(items)>4: o.write(f'  ... +{len(items)-4} more\n')
sys.stdout.buffer.write(o.getvalue().encode('utf-8','replace'))
