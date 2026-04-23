import openpyxl, math, io, sys
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

USD = 1520
SHIP_ORD1_USD = 600   # عادي
SHIP_ORD2_USD = 500   # حساس
SHIP_ORD1_IQD = SHIP_ORD1_USD * USD
SHIP_ORD2_IQD = SHIP_ORD2_USD * USD
CARTON_COST = {'S':500,'M':800,'L':1000}
BROCHURE = 400 * 200  # 80,000
UGC_CARD = 260 * 200  # 52,000
MARKETING_TOTAL = BROCHURE + UGC_CARD  # 132,000
DEFAULT_MARGIN = 0.45

# ── 1. قراءة packing list 1.9 واستخراج CBM لكل كرتون ────────────────────────
def read_packing(path, ship_iqd):
    """يقرأ packing list ويحسب شحن كل منتج بنسبة CBM"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    
    # جمع كل المنتجات مع بيانات الكرتون
    items = []
    total_cbm_declared = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 6: continue
        try: num = int(row[0])
        except: continue
        
        code = str(row[2]).strip() if row[2] else ''
        name_cn = str(row[3]).strip() if row[3] else ''
        name_en = str(row[4]).strip() if row[4] else ''
        qty = int(row[5]) if row[5] else 0
        ctn_no = row[7]  # رقم الكرتون المشترك
        unit_nw = float(row[9]) if row[9] else 0
        total_nw = float(row[10]) if row[10] else qty * unit_nw
        cbm = float(row[15]) if row[15] and row[15] != 0 else 0
        
        items.append({
            'num': num, 'code': code, 'cn': name_cn, 'en': name_en,
            'qty': qty, 'ctn_no': ctn_no, 'unit_nw': unit_nw,
            'total_nw': total_nw, 'cbm_own': cbm,
        })
    
    # حساب CBM الإجمالي من الملف
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        s = str(row[0]) if row[0] else ''
        if 'Volume' in s or 'CBM' in s.upper():
            import re
            m = re.search(r'([\d.]+)\s*CBM', s, re.IGNORECASE)
            if m: total_cbm_declared = float(m.group(1))
    
    # تجميع المنتجات حسب رقم الكرتون
    cartons = defaultdict(list)
    for item in items:
        cartons[item['ctn_no']].append(item)
    
    # حساب CBM لكل كرتون
    # الكراتين التي لها CBM مباشر: خذه
    # الكراتين المشتركة بدون CBM: وزّع CBM المتبقي بنسبة الوزن
    known_cbm = 0
    known_weight = 0
    unknown_weight = 0
    
    for ctn_no, ctn_items in cartons.items():
        ctn_cbm = max(it['cbm_own'] for it in ctn_items)
        if ctn_cbm > 0:
            known_cbm += ctn_cbm
            known_weight += sum(it['total_nw'] for it in ctn_items)
        else:
            unknown_weight += sum(it['total_nw'] for it in ctn_items)
    
    remaining_cbm = total_cbm_declared - known_cbm
    if remaining_cbm < 0: remaining_cbm = 0
    
    # توزيع CBM على كل منتج
    for ctn_no, ctn_items in cartons.items():
        ctn_cbm = max(it['cbm_own'] for it in ctn_items)
        if ctn_cbm > 0:
            # كرتون معروف الحجم → وزّع CBM بنسبة الوزن بين محتوياته
            ctn_total_nw = sum(it['total_nw'] for it in ctn_items)
            for it in ctn_items:
                it['cbm_share'] = ctn_cbm * (it['total_nw'] / ctn_total_nw) if ctn_total_nw > 0 else 0
        else:
            # كرتون مجهول الحجم → وزّع CBM المتبقي بنسبة الوزن
            for it in ctn_items:
                it['cbm_share'] = remaining_cbm * (it['total_nw'] / unknown_weight) if unknown_weight > 0 else 0
    
    # حساب تكلفة الشحن لكل قطعة بنسبة CBM
    total_cbm_calc = sum(it['cbm_share'] for it in items)
    for it in items:
        if total_cbm_calc > 0:
            it['ship_total'] = ship_iqd * (it['cbm_share'] / total_cbm_calc)
            it['ship_per_unit'] = it['ship_total'] / it['qty'] if it['qty'] > 0 else 0
        else:
            it['ship_per_unit'] = ship_iqd / sum(x['qty'] for x in items)
            it['ship_total'] = it['ship_per_unit'] * it['qty']
    
    return items, total_cbm_declared

items1, cbm1 = read_packing(
    r'C:\Users\jaafa\Downloads\ros\伊拉克-Jaafar-1.9 - packing list（普货）.xlsx', SHIP_ORD1_IQD)
items2, cbm2 = read_packing(
    r'C:\Users\jaafa\Downloads\ros\伊拉克-Jaafar-1.9 - packing list（敏货）.xlsx', SHIP_ORD2_IQD)

# ── 2. قراءة أسعار الشراء من ملف 1.5 ──────────────────────────────────────────
src = openpyxl.load_workbook(
    r'c:\Users\jaafa\Desktop\upload\FishWebClean\Yee_Products_2026_UPDATED.xlsx', data_only=True)
src_rows = list(src.active.iter_rows(values_only=True))
price_map = {}  # code → buy_usd
box_map = {}
for row in src_rows[1:]:
    if not row[0]: continue
    try:
        buy_usd = float(str(row[4]).replace(',','') or 0)
        box = str(row[5]) if row[5] else 'S'
        cn = str(row[1]) if row[1] else ''
        en = str(row[2]) if row[2] else ''
        price_map[int(row[0])] = {'usd': buy_usd, 'box': box, 'cn': cn, 'en': en}
    except: pass

# ── 3. دمج البيانات وحساب التكلفة الكاملة ─────────────────────────────────────
total_units = sum(it['qty'] for it in items1 + items2)
mkt_per_unit = round(MARKETING_TOTAL / total_units)

PRECISE_ENDS = [50,100,150,200,250,300,350,400,450,550,600,650,700,750,800,850,900,950]

def precise_price(cost, margin):
    min_p = cost / (1 - margin)
    cost_tail = int(cost) % 1000
    preferred = min(PRECISE_ENDS, key=lambda e: abs(e - cost_tail))
    base_k = int(math.ceil(min_p / 1000))
    candidate = base_k * 1000 - 1000 + preferred
    if candidate < min_p: candidate = base_k * 1000 + preferred
    if candidate < min_p: candidate += 1000
    return int(candidate)

products = []
for order_name, items in [('عادي', items1), ('حساس', items2)]:
    for it in items:
        num = it['num']
        src_data = price_map.get(num, None)
        if not src_data:
            # حاول المطابقة بالاسم الصيني
            for k, v in price_map.items():
                if v['cn'][:10] == it['cn'][:10] and it['cn']:
                    src_data = v; break
        
        buy_usd = src_data['usd'] if src_data else 0
        box = src_data['box'] if src_data else 'S'
        en = src_data['en'] if src_data else it['en']
        carton = CARTON_COST.get(box, 500)
        
        # التكلفة الكاملة = شراء + كرتون + شحن (CBM) + تسويق
        cost = buy_usd * USD + carton + it['ship_per_unit'] + mkt_per_unit
        sell = precise_price(cost, DEFAULT_MARGIN)
        margin = (sell - cost) / sell * 100 if sell > 0 else 0
        profit = sell - cost
        
        products.append({
            'num': it['num'], 'en': en[:42], 'cn': it['cn'][:20],
            'order': order_name, 'qty': it['qty'],
            'buy_usd': buy_usd, 'buy_iqd': buy_usd * USD,
            'box': box, 'carton': carton,
            'ship_unit': round(it['ship_per_unit']),
            'cbm_share': it['cbm_share'],
            'mkt': mkt_per_unit,
            'cost': round(cost), 'sell': sell,
            'margin': round(margin, 1), 'profit': round(profit),
            'total_p': round(profit * it['qty']),
            'pts': sell // 1000, 'tail': sell % 1000,
        })

total_rev = sum(p['sell']*p['qty'] for p in products)
total_cost = sum(p['cost']*p['qty'] for p in products)
total_prof = total_rev - total_cost
avg_mg = total_prof / total_rev * 100

# ── 4. بناء Excel ─────────────────────────────────────────────────────────────
def hf(c): return PatternFill('solid', fgColor=c)
def ft(bold=False, color='000000', size=9, italic=False):
    return Font(bold=bold,color=color,size=size,italic=italic,name='Calibri')
def al(h='center',v='center',wrap=False):
    return Alignment(horizontal=h,vertical=v,wrap_text=wrap,readingOrder=2)
T=Side(style='thin',color='CCCCCC')
def bd(): return Border(top=T,bottom=T,left=T,right=T)
NAVY='1A1A2E'; RED='E63946'; GOLD='FFD700'; TEAL='A8DADC'
MINT='D8F3DC'; AMB='FFF3CD'; ROSE='FFD6D6'; LGRY='F5F5F5'

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'التسعير'
ws.sheet_view.rightToLeft = True
ws.freeze_panes = 'A4'

# عنوان
ws.merge_cells('A1:T1')
c=ws['A1']
c.value='AQUAVO — تسعير 2026 | شحن موزّع بنسبة CBM (متر مكعب) | 1$=1,520د'
c.fill=hf(NAVY); c.font=ft(True,TEAL,12); c.alignment=al()
ws.row_dimensions[1].height=26

ws.merge_cells('A2:T2')
c=ws['A2']
c.value=(f'طلب1: {cbm1}CBM×{SHIP_ORD1_USD}$ | طلب2: {cbm2}CBM×{SHIP_ORD2_USD}$ | '
         f'بروشور+UGC: {MARKETING_TOTAL:,}د ({mkt_per_unit}د/قطعة) | نسبة ربح: {int(DEFAULT_MARGIN*100)}%')
c.fill=hf('16213E'); c.font=ft(False,TEAL,8,True); c.alignment=al()
ws.row_dimensions[2].height=18

HEADS=[
    ('رقم',4),('المنتج',36),('الطلب',6),('الكمية',5),
    ('شراء$',7),('شراء د',9),('كرتون',7),
    ('شحن/قطعة\n(CBM)',10),('تسويق/قطعة',9),('التكلفة\nالكاملة',11),
    ('سعر البيع',11),('هامش%',7),('ربح/قطعة',9),('ربح كلي',10),
    ('خصم5%',8),('هامش5%',7),('خصم10%',8),('هامش10%',7),
    ('نقاط',6),('ذيل',6),
]
for ci,(h,w) in enumerate(HEADS,1):
    c=ws.cell(3,ci,h)
    c.fill=hf(RED); c.font=ft(True,'FFFFFF',8); c.alignment=al(wrap=True); c.border=bd()
    ws.column_dimensions[get_column_letter(ci)].width=w
ws.row_dimensions[3].height=34

for i,p in enumerate(products):
    r=i+4; alt=LGRY if i%2==0 else 'FFFFFF'
    mg=p['margin']
    mf=MINT if mg>=48 else AMB if mg>=35 else ROSE
    s=p['sell']
    d5,d10=int(s*.95),int(s*.90)
    m5=round((d5-p['cost'])/d5*100,1) if d5>0 else 0
    m10=round((d10-p['cost'])/d10*100,1) if d10>0 else 0

    vals=[
        p['num'],p['en'],p['order'],p['qty'],
        p['buy_usd'],p['buy_iqd'],p['carton'],
        p['ship_unit'],p['mkt'],p['cost'],
        s,p['margin'],p['profit'],p['total_p'],
        d5,m5,d10,m10,
        int(p['pts']),int(p['tail']),
    ]
    for ci,v in enumerate(vals,1):
        cell=ws.cell(r,ci,v)
        cell.border=bd(); cell.font=ft(size=8)
        cell.alignment=al(h='right' if ci>2 else 'right',wrap=(ci==2))
        if ci==11: cell.fill=hf('E8F4FD'); cell.font=ft(True,'003366',10); cell.number_format='#,##0'
        elif ci==12: cell.fill=hf(mf); cell.font=ft(True,size=9)
        elif ci==8: cell.fill=hf('FFF0F0'); cell.font=ft(True,RED,8)  # شحن CBM مميز
        elif ci in(16,18): cell.fill=hf(MINT) if v>=35 else hf(AMB) if v>=20 else hf(ROSE)
        elif ci in(19,20): cell.fill=hf('FFF8E1')
        else: cell.fill=hf(alt)
    ws.row_dimensions[r].height=16

# إجماليات
tr=len(products)+4
ws.merge_cells(f'A{tr}:J{tr}')
c=ws.cell(tr,1,'الإجماليات'); c.fill=hf(NAVY); c.font=ft(True,GOLD,10); c.alignment=al(); c.border=bd()
for ci,v in [(11,int(total_rev)),(12,round(avg_mg,1)),(13,''),(14,int(total_prof))]:
    c=ws.cell(tr,ci,v); c.fill=hf(NAVY); c.font=ft(True,GOLD,10); c.alignment=al(); c.border=bd()

# ── شيت الملخص المالي ─────────────────────────────────────────────────────────
ws2=wb.create_sheet('الملخص المالي')
ws2.sheet_view.rightToLeft=True
ws2['A1']='الملخص المالي — توزيع CBM'
ws2['A1'].fill=hf(NAVY); ws2['A1'].font=ft(True,TEAL,13); ws2.merge_cells('A1:C1')
ws2.row_dimensions[1].height=26

rows2=[
    ('البند','دينار','دولار'),
    ('═══ الشحن بالـ CBM ═══','',''),
    (f'طلب 1 عادي: {cbm1} CBM',f'{SHIP_ORD1_IQD:,}',f'${SHIP_ORD1_USD}'),
    (f'طلب 2 حساس: {cbm2} CBM',f'{SHIP_ORD2_IQD:,}',f'${SHIP_ORD2_USD}'),
    (f'سعر CBM طلب1',f'{SHIP_ORD1_IQD/cbm1:,.0f} د/CBM',f'${SHIP_ORD1_USD/cbm1:.0f}/CBM'),
    (f'سعر CBM طلب2',f'{SHIP_ORD2_IQD/cbm2:,.0f} د/CBM',f'${SHIP_ORD2_USD/cbm2:.0f}/CBM'),
    ('','',''),
    ('═══ التكاليف الأخرى ═══','',''),
    ('بروشور 200×400',f'{BROCHURE:,}','—'),
    ('كارت UGC 200×260',f'{UGC_CARD:,}','—'),
    (f'تسويق/قطعة ({total_units} قطعة)',f'{mkt_per_unit}','—'),
    ('','',''),
    ('═══ النتائج ═══','',''),
    ('إجمالي الإيرادات',f'{int(total_rev):,}',f'${int(total_rev/USD):,}'),
    ('صافي الربح',f'{int(total_prof):,}',f'${int(total_prof/USD):,}'),
    ('هامش الربح',f'{avg_mg:.1f}%','—'),
    ('نقطة التعادل',f'{(SHIP_ORD1_IQD+SHIP_ORD2_IQD+MARKETING_TOTAL)/total_rev*100:.1f}%','—'),
    ('','',''),
    ('═══ الخصومات ═══','',''),
    ('ربح عند خصم 5%',f'{int(total_rev*.95-total_cost):,}','—'),
    ('ربح عند خصم 10%',f'{int(total_rev*.90-total_cost):,}','—'),
    ('ربح عند خصم 15%',f'{int(total_rev*.85-total_cost):,}','—'),
]
for ri,row3 in enumerate(rows2,2):
    for ci,v in enumerate(row3,1):
        c=ws2.cell(ri,ci,v); c.border=bd(); c.alignment=al(h='right')
        if ri==2: c.fill=hf(RED); c.font=ft(True,'FFFFFF',9)
        elif '═══' in str(row3[0]): c.fill=hf(NAVY); c.font=ft(True,GOLD,10)
        else: c.fill=hf(LGRY if ri%2==0 else 'FFFFFF'); c.font=ft(size=9)
    ws2.row_dimensions[ri].height=18
ws2.column_dimensions['A'].width=30; ws2.column_dimensions['B'].width=22; ws2.column_dimensions['C'].width=18

# حفظ
out_path=r'c:\Users\jaafa\Desktop\upload\FishWebClean\AQUAVO_تسعير_CBM_v4.xlsx'
wb.save(out_path)
o=io.StringIO()
o.write(f'تم الحفظ: {out_path}\n')
o.write(f'منتجات: {len(products)} | ايرادات: {int(total_rev):,} | ربح: {int(total_prof):,} | هامش: {avg_mg:.1f}%\n\n')
o.write('عينة — لاحظ اختلاف الشحن حسب CBM:\n')
for p in products[:6]:
    o.write(f'  #{p["num"]} {p["en"][:30]:<30} | شحن/قطعة: {p["ship_unit"]:>6,}د | تكلفة: {p["cost"]:>7,} | سعر: {p["sell"]:>7,} | هامش: {p["margin"]}%\n')
o.write('\n--- مقارنة: أحواض ثقيلة vs طعام خفيف ---\n')
heavy = [p for p in products if 'thick' in p['en'].lower() or '40X40' in p['en']]
light = [p for p in products if 'feed' in p['en'].lower() or 'food' in p['en'].lower()][:3]
for p in heavy[:3]+light:
    o.write(f'  #{p["num"]} {p["en"][:30]:<30} | شحن: {p["ship_unit"]:>6,}د\n')
sys.stdout.buffer.write(o.getvalue().encode('utf-8','replace'))
