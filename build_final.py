import openpyxl, math, io, sys
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

USD=1520; SHIP1=400; SHIP2=500; INTERNAL=58; WOOD=65; UGC_USD=70
ENDS=[50,100,150,200,250,300,350,400,450,550,600,650,700,750,800,850,900,950]
MARGINS={'حوض':0.25,'سخان':0.15,'فلتر':0.30,'إكسسوار':0.30,'تربة':0.15,
         'غذاء':0.55,'كيماوي':0.55,'اختبار':0.60,'فيتامين':0.65,'علاج':0.70,'هدية':0.15}

def get_cat(en,cn):
    s=(en+cn).lower()
    if any(k in s for k in ['feed','food','brine','worm','shrimp','饲料','粮','卵','冻干','鱼粮']): return 'غذاء'
    if any(k in s for k in ['heat','ysh','quartz','加热','武士']): return 'سخان'
    if any(k in s for k in ['filter','pump','oil film','换水','过滤','油膜','滤材','培养环','培菌','3D','16-in']): return 'فلتر'
    if any(k in s for k in ['tank','mm thick','glass','缸','6mm','8mm','5mm','stream']): return 'حوض'
    if any(k in s for k in ['spot','methylene','bactericid','白点','黄粉','亚甲基蓝']): return 'علاج'
    if any(k in s for k in ['mineral','salt','矿物盐','多维']): return 'فيتامين'
    if any(k in s for k in ['test','tester','试纸','测试']): return 'اختبار'
    if any(k in s for k in ['nitrif','stabilizer','algae','chlorine','ammonia','bacteria','descal','除','净','稳定','硝化','益生','藻','probio']): return 'كيماوي'
    if any(k in s for k in ['grass mud','水草泥']): return 'تربة'
    return 'إكسسوار'

def pprice(cost,margin):
    mp=cost/(1-margin); tail=int(cost)%1000
    pref=min(ENDS,key=lambda e:abs(e-tail))
    bk=int(math.ceil(mp/1000)); c=bk*1000-1000+pref
    if c<mp: c=bk*1000+pref
    if c<mp: c+=1000
    return int(c)

# قراءة الأسعار
prices={}
for row in openpyxl.load_workbook(r'c:\Users\jaafa\Desktop\upload\FishWebClean\客户伊拉克-Jaafar-1.5 (1).xlsx',data_only=True).active.iter_rows(values_only=True):
    code=str(row[2]).strip() if row[2] else ''
    if not code or 'Code' in code: continue
    try: prices[code]=float(str(row[6]).replace(',',''))
    except: pass

def read_pl(path):
    items=[]
    for row in openpyxl.load_workbook(path,data_only=True).active.iter_rows(values_only=True):
        try: num=int(row[0])
        except: continue
        items.append({'code':str(row[2]).strip(),'cn':str(row[3])[:25] if row[3] else '',
            'en':str(row[4])[:42] if row[4] else '','qty':int(row[5]) if row[5] else 0,
            'unw':float(row[9]) if row[9] else 0})
    return items

items1=read_pl(r'C:\Users\jaafa\Downloads\ros\伊拉克-Jaafar-1.9 - packing list（普货）.xlsx')
items2=read_pl(r'C:\Users\jaafa\Downloads\ros\伊拉克-Jaafar-1.9 - packing list（敏货）.xlsx')

# حساب القيم
tval1=sum(prices.get(i['code'],0)*i['qty'] for i in items1)
tval2=sum(prices.get(i['code'],0)*i['qty'] for i in items2)
int1=INTERNAL*tval1/(tval1+tval2); int2=INTERNAL*tval2/(tval1+tval2)
ship1_iqd=(SHIP1+int1)*USD; ship2_iqd=(SHIP2+int2)*USD
tnw1=sum(i['unw']*i['qty'] for i in items1); tnw2=sum(i['unw']*i['qty'] for i in items2)
qty1=sum(i['qty'] for i in items1); qty2=sum(i['qty'] for i in items2)

# توزيع هجين: 70% كمية + 30% وزن
qty_rate1=ship1_iqd*0.70/qty1; wt_rate1=ship1_iqd*0.30/tnw1
qty_rate2=ship2_iqd*0.70/qty2; wt_rate2=ship2_iqd*0.30/tnw2

# خشب 65$ على الأحواض فقط
tank_val=sum(prices.get(i['code'],0)*i['qty'] for i in items1 if get_cat(i['en'],i['cn'])=='حوض')
wood_per_d=WOOD/tank_val if tank_val>0 else 0
total_units=qty1+qty2; ugc_per=round(UGC_USD*USD/total_units)

products=[]
for order,items,qr,wr in [('عادي',items1,qty_rate1,wt_rate1),('حساس',items2,qty_rate2,wt_rate2)]:
    for it in items:
        up=prices.get(it['code'],0); cat=get_cat(it['en'],it['cn'])
        ship_u=round(qr+it['unw']*wr)
        wood_u=round(up*wood_per_d*USD) if(cat=='حوض' and order=='عادي') else 0
        if cat == 'حوض': carton = 0
        elif it['unw'] < 0.4: carton = 150
        elif it['unw'] <= 2.0: carton = 300
        else: carton = 500
        cost=round(up*USD+ship_u+wood_u+carton+ugc_per)
        margin=MARGINS.get(cat,0.30)
        sell=pprice(cost,margin)
        mg=round((sell-cost)/sell*100,1) if sell>0 else 0
        products.append({'en':it['en'],'cn':it['cn'],'order':order,'cat':cat,
            'qty':it['qty'],'buy':up,'buy_iqd':round(up*USD),'unw':it['unw'],
            'carton':carton,'ship':ship_u,'wood':wood_u,'ugc':ugc_per,
            'cost':cost,'sell':sell,'margin':mg,'target':round(margin*100),
            'profit':sell-cost,'total_p':(sell-cost)*it['qty']})

trev=sum(p['sell']*p['qty'] for p in products)
tcost=sum(p['cost']*p['qty'] for p in products)
tprof=trev-tcost; amg=tprof/trev*100

# Excel
def hf(c): return PatternFill('solid',fgColor=c)
def ft(b=False,c='000000',s=9,i=False): return Font(bold=b,color=c,size=s,italic=i,name='Calibri')
def al(h='center',v='center',w=False): return Alignment(horizontal=h,vertical=v,wrap_text=w,readingOrder=2)
T=Side(style='thin',color='CCCCCC')
def bd(): return Border(top=T,bottom=T,left=T,right=T)
NV='1A1A2E';RD='E63946';GD='FFD700';TL='A8DADC';MT='D8F3DC';AM='FFF3CD';RS='FFD6D6';LG='F5F5F5'
wb=openpyxl.Workbook(); ws=wb.active; ws.title='التسعير'; ws.sheet_view.rightToLeft=True; ws.freeze_panes='A4'
ws.merge_cells('A1:R1'); ws['A1'].value='AQUAVO — التسعير النهائي 2026 | شحن هجين (70%كمية+30%وزن) | 1$=1,520د'
ws['A1'].fill=hf(NV); ws['A1'].font=ft(True,TL,11); ws['A1'].alignment=al()
ws.merge_cells('A2:R2')
ws['A2'].value=f'حوض:25% سخان/تربة:15% فلتر:30% غذاء:55% كيماوي:55% اختبار:60% فيتامين:65% علاج:70%'
ws['A2'].fill=hf('16213E'); ws['A2'].font=ft(False,TL,8,True); ws['A2'].alignment=al()
H=[('م',3),('المنتج',33),('الفئة',7),('كمية',4),('شراء$',6),('شراء د',8),
   ('شحن/ق',7),('خشب',5),('كرتون',5),('التكلفة',9),('سعر بيع',9),
   ('هامش%',6),('ربح/ق',8),('ربح كلي',9),('خصم5%',6),('خصم10%',6),('نقاط',5),('ذيل',4)]
for ci,(h,w) in enumerate(H,1):
    c=ws.cell(3,ci,h); c.fill=hf(RD); c.font=ft(True,'FFFFFF',7); c.alignment=al(w=True); c.border=bd()
    ws.column_dimensions[get_column_letter(ci)].width=w
for i,p in enumerate(products):
    r=i+4; bg=LG if i%2==0 else 'FFFFFF'
    mf=MT if p['margin']>=p['target'] else AM if p['margin']>=p['target']-5 else RS
    s=p['sell']
    d5=round((s*.95-p['cost'])/(s*.95)*100,1) if s>0 else 0
    d10=round((s*.90-p['cost'])/(s*.90)*100,1) if s>0 else 0
    vals=[i+1,p['en'],p['cat'],p['qty'],round(p['buy'],2),p['buy_iqd'],
          p['ship'],p['wood'],p['carton'],p['cost'],s,p['margin'],p['profit'],p['total_p'],d5,d10,s//1000,s%1000]
    for ci,v in enumerate(vals,1):
        cell=ws.cell(r,ci,v); cell.border=bd(); cell.font=ft(s=7); cell.alignment=al(h='right',w=(ci==2))
        if ci==11: cell.fill=hf('E8F4FD'); cell.font=ft(True,'003366',9); cell.number_format='#,##0'
        elif ci==3: cell.fill=hf('FFF0E0'); cell.font=ft(True,s=7)
        elif ci==12: cell.fill=hf(mf); cell.font=ft(True,s=8)
        else: cell.fill=hf(bg)
tr=len(products)+4
ws.merge_cells(f'A{tr}:J{tr}')
c=ws.cell(tr,1,'الإجماليات'); c.fill=hf(NV); c.font=ft(True,GD,10); c.alignment=al(); c.border=bd()
for ci,v in [(11,int(trev)),(12,round(amg,1)),(14,int(tprof))]:
    c=ws.cell(tr,ci,v); c.fill=hf(NV); c.font=ft(True,GD,9); c.alignment=al(); c.border=bd()

out=r'c:\Users\jaafa\Desktop\upload\FishWebClean\AQUAVO_FINAL_v10.xlsx'
wb.save(out)
o=io.StringIO()
o.write(f'SAVED: {out}\n{len(products)} products | Rev:{int(trev):,} | Profit:{int(tprof):,} | Margin:{amg:.1f}%\n\n')
from collections import defaultdict
by_cat=defaultdict(list)
for p in products: by_cat[p['cat']].append(p)
for cat in ['حوض','سخان','تربة','فلتر','غذاء','علاج','كيماوي','اختبار','فيتامين','إكسسوار']:
    if cat not in by_cat: continue
    for p in by_cat[cat][:3]:
        o.write(f'{cat:>8} {p["en"][:28]:<28} cost:{p["cost"]:>6} sell:{p["sell"]:>6} mg:{p["margin"]}%\n')
    if len(by_cat[cat])>3: o.write(f'         ... +{len(by_cat[cat])-3}\n')
sys.stdout.buffer.write(o.getvalue().encode('utf-8','replace'))
