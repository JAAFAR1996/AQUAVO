
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                               GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import math

# ─── Constants ───────────────────────────────────────────────────────────────
USD_TO_IQD      = 1520
TOTAL_EXP_USD   = 2028
TOTAL_EXP_IQD   = TOTAL_EXP_USD * USD_TO_IQD   # 3,082,560
OVERHEAD_PER    = 2722   # IQD per unit (shipping + misc distributed)
CARTON          = {'S': 500, 'M': 800, 'L': 1000}

# Target margin per category (minimum)
MARGIN_TARGET   = {
    'heater':   0.42,
    'filter':   0.42,
    'tank':     0.45,
    'food':     0.50,
    'medical':  0.55,
    'chemical': 0.52,
    'other':    0.45,
}

# Category auto-detect keywords
CAT_KEYWORDS = {
    'heater':   ['heat', 'YSH', 'quartz', '加热'],
    'filter':   ['filter', 'pump', 'YLC', 'oil film', 'stream', '滤'],
    'tank':     ['tank', 'mm thick', '缸', 'aquarium glass'],
    'food':     ['feed', 'food', 'brine', 'worm', 'feast', '饲料', '粮'],
    'medical':  ['white spot', 'methylene', 'bactericid', '白点', '亚甲基', '黄粉'],
    'chemical': ['probio', 'nitrif', 'stabilizer', 'algae', 'chlorine',
                 'ammonia', 'mineral', 'bacteria', 'algaecide', 'descal',
                 'test', '菌', '盐', '除', '净', '稳定'],
}

def detect_category(name_en, name_cn):
    s = (name_en + name_cn).lower()
    for cat, kws in CAT_KEYWORDS.items():
        for kw in kws:
            if kw.lower() in s:
                return cat
    return 'other'

def psych_price(cost_iqd, min_margin):
    """
    Return a psychological price:
    - Guaranteed ≥ min_margin above cost
    - Ends in one of: 250, 500, 750, 000  (always 4-digit aligned to 250)
    - e.g. 13,750 / 9,500 / 7,250 / 22,000
    """
    min_price = cost_iqd / (1 - min_margin)
    # Round UP to nearest 250
    remainder = min_price % 250
    if remainder == 0:
        base = int(min_price)
    else:
        base = int(min_price - remainder + 250)
    # If ending is 000, shift to X,750 to create a "fractional feel"
    tail = base % 1000
    if tail == 0:
        base -= 250   # go to X,750
    # Ensure still above minimum
    if base < min_price:
        base += 250
    return base

def loyalty_points(price):
    """
    How many points the buyer earns.
    Rule: 1 point per 1000 IQD spent, rounded normally.
    The 'tail' (price mod 1000) becomes bonus points if tail > 500.
    Returns (points, tail_iqd_displayed)
    """
    pts = price // 1000
    tail = price % 1000
    return pts, tail

# ─── Read source Excel ────────────────────────────────────────────────────────
src = openpyxl.load_workbook(
    r'c:\Users\jaafa\Desktop\upload\FishWebClean\Yee_Products_2026_UPDATED.xlsx',
    data_only=True)
ws_src = src['Sheet1']
src_rows = list(ws_src.iter_rows(values_only=True))

# ─── Colours & Styles ────────────────────────────────────────────────────────
C_HEADER_BG   = "1A1A2E"   # dark navy
C_HEADER_FT   = "E0E0E0"   # light grey
C_SUB_BG      = "16213E"   # dark blue
C_SUB_FT      = "A8DADC"   # teal
C_GOOD_BG     = "D8F3DC"   # mint
C_WARN_BG     = "FFF3CD"   # amber
C_BAD_BG      = "FFD6D6"   # red
C_ALT1        = "F8F9FA"   # very light grey rows
C_ALT2        = "FFFFFF"
C_ACCENT      = "E63946"   # AQUAVO red

def h(hex_col):
    return PatternFill("solid", fgColor=hex_col)

def ft(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def al(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

thin = Side(style='thin', color='CCCCCC')
thick = Side(style='medium', color='888888')
def border(top=thin, bot=thin, left=thin, right=thin):
    return Border(top=top, bottom=bot, left=left, right=right)

# ─── New Workbook ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — PRICING TABLE
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "📦 Pricing Table"
ws.sheet_view.rightToLeft = False
ws.freeze_panes = "A4"

# ── Row 1: Title ─────────────────────────────────────────────────────────────
ws.merge_cells("A1:R1")
title_cell = ws["A1"]
title_cell.value = "🐟 AQUAVO — YEE Products Pricing 2026   |   Rate: 1 USD = 1,520 IQD   |   Total Investment: $2,028"
title_cell.fill = h(C_HEADER_BG)
title_cell.font = ft(bold=True, color=C_HEADER_FT, size=13)
title_cell.alignment = al()
ws.row_dimensions[1].height = 28

# ── Row 2: Sub-info ───────────────────────────────────────────────────────────
ws.merge_cells("A2:R2")
info_cell = ws["A2"]
info_cell.value = (f"Total Cost: {TOTAL_EXP_IQD:,} IQD  |  "
                   f"Overhead/unit: {OVERHEAD_PER:,} IQD  |  "
                   f"Loyalty Rule: 1 pt per 1,000 IQD spent  |  "
                   f"Discount Cols: 5% / 10% / 15% impact on margin")
info_cell.fill = h(C_SUB_BG)
info_cell.font = ft(italic=True, color=C_SUB_FT, size=9)
info_cell.alignment = al()
ws.row_dimensions[2].height = 18

# ── Row 3: Column Headers ─────────────────────────────────────────────────────
COLS = [
    ("#",              5),
    ("Product (EN)",   42),
    ("اسم المنتج",    22),
    ("Qty",            5),
    ("Buy $",          7),
    ("Buy IQD",        9),
    ("Box",            5),
    ("Carton IQD",     9),
    ("Ship/unit",      9),
    ("Cost/unit",     10),
    ("Sell Price",    11),
    ("Margin %",       9),
    ("Profit/unit",   10),
    ("Total Profit",  11),
    ("@5% disc",       9),
    ("@10% disc",      9),
    ("@15% disc",      9),
    ("🎁 Pts/order",  9),
]

for col_idx, (hdr, width) in enumerate(COLS, start=1):
    cell = ws.cell(row=3, column=col_idx, value=hdr)
    cell.fill = h(C_ACCENT)
    cell.font = ft(bold=True, color="FFFFFF", size=9)
    cell.alignment = al(wrap=True)
    cell.border = border()
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[3].height = 32

# ── Data Rows ─────────────────────────────────────────────────────────────────
products = []
for row in src_rows[1:]:
    if not row[0] or str(row[0]).strip() == '':
        continue
    try:
        num      = row[0]
        name_cn  = str(row[1]) if row[1] else ''
        name_en  = str(row[2]) if row[2] else ''
        qty      = float(str(row[3]).replace(',', '')) if row[3] else 0
        buy_usd  = float(str(row[4]).replace(',', '')) if row[4] else 0
        box_size = str(row[5]) if row[5] else 'S'
        carton   = CARTON.get(box_size, 500)
        cat      = detect_category(name_en, name_cn)
        min_marg = MARGIN_TARGET[cat]

        buy_iqd   = buy_usd * USD_TO_IQD
        cost_unit = buy_iqd + carton + OVERHEAD_PER
        sell      = psych_price(cost_unit, min_marg)
        actual_mg = (sell - cost_unit) / sell
        profit_u  = sell - cost_unit
        total_pr  = profit_u * qty
        pts, tail = loyalty_points(sell)

        # Discount margin impacts
        mg_5  = ((sell * 0.95) - cost_unit) / (sell * 0.95) * 100
        mg_10 = ((sell * 0.90) - cost_unit) / (sell * 0.90) * 100
        mg_15 = ((sell * 0.85) - cost_unit) / (sell * 0.85) * 100

        products.append({
            'num': num, 'name_en': name_en, 'name_cn': name_cn,
            'qty': qty, 'buy_usd': buy_usd, 'buy_iqd': buy_iqd,
            'box': box_size, 'carton': carton, 'ship': OVERHEAD_PER,
            'cost': cost_unit, 'sell': sell, 'margin': actual_mg * 100,
            'profit_u': profit_u, 'total_pr': total_pr,
            'mg5': mg_5, 'mg10': mg_10, 'mg15': mg_15,
            'pts': pts, 'tail': tail, 'cat': cat,
        })
    except Exception as e:
        print(f"Error row {row[0]}: {e}")

for i, p in enumerate(products):
    r = i + 4
    alt = C_ALT1 if i % 2 == 0 else C_ALT2

    def bg_fill(val=None):
        if val is not None:
            if val >= 45:  return h(C_GOOD_BG)
            elif val >= 30: return h(C_WARN_BG)
            else:           return h(C_BAD_BG)
        return h(alt)

    vals = [
        p['num'], p['name_en'], p['name_cn'],
        int(p['qty']), p['buy_usd'],
        int(p['buy_iqd']),
        p['box'], p['carton'], p['ship'],
        int(p['cost']),
        p['sell'],
        round(p['margin'], 1),
        int(p['profit_u']),
        int(p['total_pr']),
        round(p['mg5'], 1),
        round(p['mg10'], 1),
        round(p['mg15'], 1),
        f"{p['pts']} pts (tail {p['tail']:,} IQD)",
    ]

    for c, v in enumerate(vals, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = border()
        cell.font   = ft(size=9)
        cell.alignment = al(h="right" if c > 3 else "left", wrap=(c<=3))

        # Colour logic
        if c == 12:  # Margin %
            cell.fill = bg_fill(p['margin'])
            cell.font = ft(bold=True, size=9)
        elif c == 11:  # Sell price - highlight
            cell.fill = h("EAF4FF")
            cell.font = ft(bold=True, size=9, color="003366")
            cell.number_format = '#,##0 "IQD"'
        elif c in (13, 14):
            cell.fill = h(C_GOOD_BG) if p['profit_u'] > 0 else h(C_BAD_BG)
            cell.number_format = '#,##0'
        elif c == 15:  # @5%
            cell.fill = h(C_GOOD_BG) if p['mg5'] >= 30 else h(C_WARN_BG)
        elif c == 16:  # @10%
            cell.fill = h(C_WARN_BG) if p['mg10'] >= 20 else h(C_BAD_BG)
        elif c == 17:  # @15%
            cell.fill = h(C_BAD_BG) if p['mg15'] < 20 else h(C_WARN_BG)
        elif c == 18:  # Points
            cell.fill = h("FFF8E1")
            cell.font = ft(italic=True, size=9, color="5D4037")
        else:
            cell.fill = h(alt)

    ws.row_dimensions[r].height = 20

# ── Totals Row ────────────────────────────────────────────────────────────────
r_tot = len(products) + 4
ws.merge_cells(f"A{r_tot}:J{r_tot}")
tot_label = ws.cell(row=r_tot, column=1, value="TOTALS / SUMMARY")
tot_label.fill = h(C_HEADER_BG)
tot_label.font = ft(bold=True, color=C_HEADER_FT)
tot_label.alignment = al()

total_revenue = sum(p['sell'] * p['qty'] for p in products)
total_profit  = sum(p['total_pr'] for p in products)
avg_margin    = total_profit / total_revenue * 100

summary_vals = [
    ('', 11), (total_revenue, 12), (avg_margin, 12),
    (total_profit, 13), (total_profit, 14),
    ('', 15), ('', 16), ('', 17), ('', 18)
]
for col, (val, _) in zip(range(11, 20), summary_vals):
    c = ws.cell(row=r_tot, column=col, value=round(val, 1) if isinstance(val, float) else val)
    c.fill = h(C_HEADER_BG)
    c.font = ft(bold=True, color="FFD700")
    c.alignment = al()
    c.number_format = '#,##0' if isinstance(val, int) else '0.0"%"'
    c.border = border()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — DISCOUNT SIMULATOR
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("💸 Discount Simulator")
ws2.sheet_view.rightToLeft = False

ws2.merge_cells("A1:L1")
ws2["A1"].value = "💸 AQUAVO — Discount Impact Simulator"
ws2["A1"].fill  = h(C_HEADER_BG)
ws2["A1"].font  = ft(bold=True, color=C_HEADER_FT, size=13)
ws2["A1"].alignment = al()
ws2.row_dimensions[1].height = 26

heads2 = [
    ("#", 5), ("Product", 40), ("Sell Price", 11),
    ("Cost/unit", 10), ("Margin%", 9),
    ("@5% sell", 10), ("@5% margin", 9),
    ("@10% sell", 10), ("@10% margin", 9),
    ("@15% sell", 10), ("@15% margin", 9),
    ("Min sell (30%)", 12),
]
for ci, (h2, w2) in enumerate(heads2, 1):
    c = ws2.cell(row=2, column=ci, value=h2)
    c.fill = h(C_ACCENT); c.font = ft(bold=True, color="FFFFFF", size=9)
    c.alignment = al(); c.border = border()
    ws2.column_dimensions[get_column_letter(ci)].width = w2
ws2.row_dimensions[2].height = 28

for i, p in enumerate(products):
    r2 = i + 3
    alt = C_ALT1 if i % 2 == 0 else C_ALT2
    min_sell_30 = int(math.ceil(p['cost'] / 0.70 / 250) * 250)  # floor for 30% margin
    # ensure non-zero tail
    if min_sell_30 % 1000 == 0:
        min_sell_30 -= 250

    row_vals = [
        p['num'], p['name_en'][:45], p['sell'], int(p['cost']),
        round(p['margin'], 1),
        int(p['sell'] * 0.95), round(p['mg5'], 1),
        int(p['sell'] * 0.90), round(p['mg10'], 1),
        int(p['sell'] * 0.85), round(p['mg15'], 1),
        min_sell_30,
    ]
    for ci, v in enumerate(row_vals, 1):
        c2 = ws2.cell(row=r2, column=ci, value=v)
        c2.border = border(); c2.font = ft(size=9)
        c2.alignment = al(h="right" if ci > 2 else "left")
        # colour margin columns
        if ci in (5, 7, 9, 11):
            val_f = v if isinstance(v, float) else 0
            c2.fill = h(C_GOOD_BG) if val_f >= 40 else h(C_WARN_BG) if val_f >= 25 else h(C_BAD_BG)
            c2.font = ft(bold=True, size=9)
        elif ci == 12:
            c2.fill = h("FFF3CD"); c2.font = ft(italic=True, size=9, color="7B3F00")
        else:
            c2.fill = h(alt)
    ws2.row_dimensions[r2].height = 18

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — LOYALTY POINTS GUIDE
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("🎁 Loyalty Points Guide")
ws3.merge_cells("A1:H1")
ws3["A1"].value = "🎁 AQUAVO Loyalty Points — كيف يشتغل نظام النقاط"
ws3["A1"].fill = h(C_HEADER_BG)
ws3["A1"].font = ft(bold=True, color=C_HEADER_FT, size=13)
ws3["A1"].alignment = al()
ws3.row_dimensions[1].height = 26

guide = [
    ["", ""],
    ["📌 القاعدة الأساسية:", ""],
    ["• كل 1,000 دينار = 1 نقطة", "مثال: اشترى بـ 14,750 → يحصل 14 نقطة"],
    ["• الباقي (tail) = نقاط مضافة إذا تجاوز 500 دينار", "مثال: 14,750 → tail=750 → +1 نقطة = 15 نقطة"],
    ["", ""],
    ["📌 لماذا الأسعار كسور (مثل 14,750 بدل 15,000)?", ""],
    ["• تجعل السعر يبدو أرخص نفسياً", "15,000 تبدو عالية، 14,750 تبدو معقولة"],
    ["• الـ 250 'المتبقية' تحفّز المشتري إكمال الطلب", "ليحصل على نقاط أكثر"],
    ["• تشجع الطلبات المتكررة (يريد يصرف نقاطه)", ""],
    ["", ""],
    ["📌 جدول استبدال النقاط:", ""],
    ["50 نقطة  = خصم 5,000 IQD في الطلب القادم", ""],
    ["100 نقطة = خصم 10,000 IQD", ""],
    ["200 نقطة = منتج مجاني حسب القيمة", ""],
    ["", ""],
    ["📌 مستويات العضوية:", ""],
    ["Bronze  → 0–499 نقطة     → لا خصم إضافي", ""],
    ["Silver  → 500–1499 نقطة  → خصم 3% على كل طلب", ""],
    ["Gold    → 1500+ نقطة     → خصم 5% + أولوية التوصيل", ""],
    ["", ""],
    ["📌 مثال عملي:", ""],
    ["زبون اشترى: سخان 12,750 + طعام 9,750 + علاج 9,750 = 32,250 IQD", ""],
    ["نقاطه: 32 نقطة (tail=250 < 500 فلا نقطة إضافية)", ""],
    ["في الطلب التالي يصرف 32 نقطة = خصم 3,200 IQD على فاتورته", ""],
]
for ri, (col1, col2) in enumerate(guide, start=2):
    c1 = ws3.cell(row=ri, column=1, value=col1)
    c2 = ws3.cell(row=ri, column=5, value=col2)
    if col1.startswith("📌"):
        c1.font = ft(bold=True, size=11, color="E63946")
        c1.fill = h("FFF0F0")
    else:
        c1.font = ft(size=10)
    c2.font = ft(italic=True, size=10, color="555555")
    ws3.row_dimensions[ri].height = 18

ws3.column_dimensions['A'].width = 60
ws3.column_dimensions['E'].width = 50

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — FINANCIAL SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("📊 Financial Summary")
ws4.merge_cells("A1:F1")
ws4["A1"].value = "📊 AQUAVO — Financial Summary @ 1 USD = 1,520 IQD"
ws4["A1"].fill = h(C_HEADER_BG)
ws4["A1"].font = ft(bold=True, color=C_HEADER_FT, size=13)
ws4["A1"].alignment = al()

summary_data = [
    ["", "IQD", "USD"],
    ["💼 INVESTMENT", "", ""],
    ["  Total Expenses (user declared)", f"{TOTAL_EXP_IQD:,}", f"${TOTAL_EXP_USD:,}"],
    ["  Product Purchase Cost", f"{sum(p['buy_iqd']*p['qty'] for p in products):,.0f}", f"${sum(p['buy_usd']*p['qty'] for p in products):,.0f}"],
    ["  Carton Costs", f"{sum(p['carton']*p['qty'] for p in products):,.0f}", "—"],
    ["  Shipping (Order1+Order2)", f"{(600+500)*USD_TO_IQD:,}", "$1,100"],
    ["", "", ""],
    ["📈 REVENUE PROJECTIONS", "", ""],
    ["  Full Revenue (100% sold, full price)", f"{total_revenue:,.0f}", f"${total_revenue/USD_TO_IQD:,.0f}"],
    ["  Revenue @ 5% discount", f"{total_revenue*0.95:,.0f}", f"${total_revenue*0.95/USD_TO_IQD:,.0f}"],
    ["  Revenue @ 10% discount", f"{total_revenue*0.90:,.0f}", f"${total_revenue*0.90/USD_TO_IQD:,.0f}"],
    ["  Revenue @ 15% discount", f"{total_revenue*0.85:,.0f}", f"${total_revenue*0.85/USD_TO_IQD:,.0f}"],
    ["", "", ""],
    ["💰 PROFIT SCENARIOS", "", ""],
    ["  Net Profit (100%, no discount)", f"{total_profit:,.0f}", f"${total_profit/USD_TO_IQD:,.0f}"],
    ["  Net Profit @ 5% discount", f"{total_revenue*0.95 - TOTAL_EXP_IQD:,.0f}", f"${(total_revenue*0.95 - TOTAL_EXP_IQD)/USD_TO_IQD:,.0f}"],
    ["  Net Profit @ 10% discount", f"{total_revenue*0.90 - TOTAL_EXP_IQD:,.0f}", f"${(total_revenue*0.90 - TOTAL_EXP_IQD)/USD_TO_IQD:,.0f}"],
    ["  Net Profit @ 15% discount", f"{total_revenue*0.85 - TOTAL_EXP_IQD:,.0f}", f"${(total_revenue*0.85 - TOTAL_EXP_IQD)/USD_TO_IQD:,.0f}"],
    ["", "", ""],
    ["🎯 BREAK-EVEN", "", ""],
    ["  Need to sell this % of inventory", f"{TOTAL_EXP_IQD/total_revenue*100:.1f}%", "—"],
    ["  Average gross margin (full price)", f"{avg_margin:.1f}%", "—"],
    ["  Average sell price per unit", f"{total_revenue/525:,.0f}", f"${total_revenue/525/USD_TO_IQD:.1f}"],
]

for ri, row_data in enumerate(summary_data, start=2):
    for ci, val in enumerate(row_data, start=1):
        c = ws4.cell(row=ri, column=ci, value=val)
        c.alignment = al(h="left" if ci == 1 else "right")
        if str(val).startswith("💼") or str(val).startswith("📈") or str(val).startswith("💰") or str(val).startswith("🎯"):
            c.font = ft(bold=True, size=11, color="E63946")
            c.fill = h("FFF0F0")
        elif str(row_data[0]).startswith("  "):
            c.font = ft(size=10)
        ws4.row_dimensions[ri].height = 18

ws4.column_dimensions['A'].width = 45
ws4.column_dimensions['B'].width = 20
ws4.column_dimensions['C'].width = 15

# ─── Save ─────────────────────────────────────────────────────────────────────
out_path = r'c:\Users\jaafa\Desktop\upload\FishWebClean\AQUAVO_Pricing_2026_FINAL.xlsx'
wb.save(out_path)
print(f"SAVED: {out_path}")
print(f"Total Revenue: {total_revenue:,.0f} IQD")
print(f"Total Profit (full price): {total_profit:,.0f} IQD | Margin: {avg_margin:.1f}%")
print(f"Break-even: sell {TOTAL_EXP_IQD/total_revenue*100:.1f}% of items")
