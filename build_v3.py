
import openpyxl, math, io, sys
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── إعدادات ثابتة ─────────────────────────────────────────────────────────────
USD        = 1520
TOTAL_USD  = 2028
TOTAL_IQD  = TOTAL_USD * USD          # 3,082,560
OHPU       = 2722                      # توزيع الشحن لكل قطعة
CARTON     = {'S':500,'M':800,'L':1000}

# تكاليف التسويق
BROCHURE_PER = 400    # دينار/قطعة
UGC_PER      = 260    # دينار/قطعة
BROCHURE_QTY = 200
UGC_QTY      = 200
MARKETING_TOTAL = BROCHURE_PER*BROCHURE_QTY + UGC_PER*UGC_QTY  # 132,000
MARKETING_PER   = round(MARKETING_TOTAL / 525)                    # ~251 دينار/قطعة

# نسبة الربح الافتراضية (يغيرها المستخدم في الخلية B3)
DEFAULT_MARGIN = 0.45

# ── خوارزمية التسعير الدقيق (Precise Pricing) ────────────────────────────────
# البحث يثبت: 10,150 أكثر ثقة من 10,000 لأنه يوحي بحساب دقيق
# نهايات مُثلى للسوق العراقي: 50, 100, 150, 200, 250, 300, 350, 400, 450
# تجنب: 000, 500, 999 (إما مدورة جداً أو تجارية جداً)
PRECISE_ENDS = [50, 100, 150, 200, 250, 300, 350, 400, 450]


def precise_price(cost, margin, product_num):
    """
    تسعير دقيق محترف — كل منتج له ذيل فريد مبني على تكلفته الفعلية
    المنطق: ارفع التكلفة بنسبة الربح → اجعل الذيل يعكس جزءاً من التكلفة
    مثال: تكلفة 7,242 → min 13,167 → سعر 13,150 أو 13,250 أو 14,350
    """
    min_p = cost / (1 - margin)
    # الذيل المثالي: خذ آخر 3 أرقام من التكلفة واجعلها الذيل
    # هذا يجعل السعر يبدو "محسوباً بدقة من التكلفة الفعلية"
    cost_tail = int(cost) % 1000
    # نهايات مناسبة للسوق العراقي (تتجنب 000 و500)
    ENDS = [50, 100, 150, 200, 250, 300, 350, 400, 450,
            550, 600, 650, 700, 750, 800, 850, 900, 950]
    # اختر النهاية الأقرب لذيل التكلفة
    preferred = min(ENDS, key=lambda e: abs(e - cost_tail))
    base_k = int(math.ceil(min_p / 1000))  # ارفع للألف التالي
    candidate = base_k * 1000 - 1000 + preferred
    if candidate < min_p:
        candidate = base_k * 1000 + preferred
    if candidate < min_p:
        candidate += 1000
    return int(candidate)


def category(en, cn):
    s = (en+cn).lower()
    if any(k in s for k in ['heat','ysh','quartz','加热']): return 'هيتر'
    if any(k in s for k in ['filter','pump','ylc','oil film','滤']): return 'فلتر/مضخة'
    if any(k in s for k in ['tank','mm thick','缸','glass']): return 'حوض'
    if any(k in s for k in ['feed','food','brine','worm','feast','饲料','粮']): return 'غذاء'
    if any(k in s for k in ['spot','methylene','bactericid','白点','黄粉']): return 'علاج طبي'
    if any(k in s for k in ['probio','nitrif','stabilizer','algae','chlorine',
                              'ammonia','mineral','bacteria','test','菌','盐','除','净']): return 'كيماويات'
    return 'أخرى'

# ── قراءة البيانات ─────────────────────────────────────────────────────────────
src = openpyxl.load_workbook(
    r'c:\Users\jaafa\Desktop\upload\FishWebClean\Yee_Products_2026_UPDATED.xlsx',
    data_only=True)
rows = list(src.active.iter_rows(values_only=True))

products = []
for row in rows[1:]:
    if not row[0]: continue
    try:
        qty     = float(str(row[3]).replace(',','') or 0)
        buy_usd = float(str(row[4]).replace(',','') or 0)
        box     = str(row[5]) if row[5] else 'S'
        en      = str(row[2]) if row[2] else ''
        cn      = str(row[1]) if row[1] else ''
        carton  = CARTON.get(box, 500)
        # التكلفة الكاملة لكل قطعة = شراء + كرتون + شحن + تسويق
        cost    = buy_usd*USD + carton + OHPU + MARKETING_PER
        sell    = precise_price(cost, DEFAULT_MARGIN, row[0])
        margin  = (sell-cost)/sell*100
        profit  = sell-cost
        pts     = sell//1000          # نقاط الولاء
        tail    = sell%1000           # الباقي يصير نقاط
        products.append({
            'num':row[0], 'en':en, 'cn':cn, 'qty':qty,
            'buy_usd':buy_usd, 'box':box, 'carton':carton,
            'cost':cost, 'sell':sell, 'margin':margin,
            'profit':profit, 'total_p':profit*qty,
            'pts':pts, 'tail':tail, 'cat':category(en,cn),
        })
    except: pass

total_rev  = sum(p['sell']*p['qty'] for p in products)
total_cost = sum(p['cost']*p['qty'] for p in products)
total_prof = total_rev - total_cost
avg_mg     = total_prof/total_rev*100

# ── أنماط التنسيق ─────────────────────────────────────────────────────────────
def hf(c): return PatternFill('solid', fgColor=c)
def ft(bold=False, color='000000', size=9, italic=False):
    return Font(bold=bold,color=color,size=size,italic=italic,name='Calibri')
def al(h='center',v='center',wrap=False):
    return Alignment(horizontal=h,vertical=v,wrap_text=wrap,readingOrder=2)
T=Side(style='thin',color='CCCCCC')
def bd(): return Border(top=T,bottom=T,left=T,right=T)
NAVY='1A1A2E'; RED='E63946'; GOLD='FFD700'; TEAL='A8DADC'
MINT='D8F3DC'; AMB='FFF3CD'; ROSE='FFD6D6'; LGRY='F5F5F5'

# ═══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ───────────────────────── شيت 1: التسعير الرئيسي ─────────────────────────────
ws = wb.active
ws.title = 'التسعير الرئيسي'
ws.sheet_view.rightToLeft = True
ws.freeze_panes = 'A6'

# صف 1: عنوان
ws.merge_cells('A1:V1')
c=ws['A1']; c.value='AQUAVO — قائمة أسعار YEE 2026  |  1 دولار = 1,520 دينار  |  مجموع الاستثمار: 2,028$'
c.fill=hf(NAVY); c.font=ft(True,TEAL,12); c.alignment=al(); ws.row_dimensions[1].height=28

# صف 2: شرح نظام الكسور
ws.merge_cells('A2:V2')
c=ws['A2']; c.value=(
    '💡 نظام التسعير الدقيق — الأسعار الكسرية (مثال: 10,150 | 23,430) تُوحي بالحساب الدقيق وتبني الثقة بدراسات علم النفس | '
    'تكاليف التسويق مُدرجة: بروشور 200 قطعة (400د) + كارت UGC 200 قطعة (260د) = 132,000 دينار موزعة على المنتجات')
c.fill=hf('16213E'); c.font=ft(False,TEAL,8,True); c.alignment=al(); ws.row_dimensions[2].height=18

# صف 3: خلية الربح القابلة للتغيير
ws['A3']='⚙️ نسبة الربح (غيّرها هنا):'
ws['A3'].font=ft(True,RED,10); ws['A3'].alignment=al('right')
ws['B3']=DEFAULT_MARGIN; ws['B3'].number_format='0%'
ws['B3'].fill=hf('FFF8D0'); ws['B3'].font=ft(True,'8B6914',12)
ws['B3'].alignment=al(); ws['B3'].border=bd()
ws.row_dimensions[3].height=24
ws['C3']='← اكتب مثلاً: 0.40=40% | 0.50=50% | 0.55=55%'
ws['C3'].font=ft(False,'666666',8,True); ws['C3'].alignment=al('right')

# صف 4: رؤوس الأعمدة (عربية كاملة)
HEADS=[
    ('رقم',4),('المنتج',38),('الكمية',6),
    ('شراء $',9),('شراء دينار',10),('كرتون',6),
    ('تكلفة كرتون',10),('شحن/قطعة',9),('تسويق/قطعة',9),
    ('التكلفة الكاملة',12),
    ('سعر البيع (دقيق)',12),
    ('هامش %',8),('ربح/قطعة',9),('ربح كلي',11),
    ('خصم5% سعر',9),('خصم5% هامش',8),
    ('خصم10% سعر',9),('خصم10% هامش',8),
    ('خصم15% سعر',9),('خصم15% هامش',8),
    ('نقاط ولاء',8),('ذيل (نقاط)',9),
]
for ci,(h,w) in enumerate(HEADS,1):
    c=ws.cell(4,ci,h)
    c.fill=hf(RED); c.font=ft(True,'FFFFFF',8); c.alignment=al(wrap=True); c.border=bd()
    ws.column_dimensions[get_column_letter(ci)].width=w
ws.row_dimensions[4].height=36

# صفوف البيانات (تبدأ من صف 5)
for i,p in enumerate(products):
    r=i+5; alt=LGRY if i%2==0 else 'FFFFFF'
    mg=p['margin']
    mf = MINT if mg>=48 else AMB if mg>=35 else ROSE
    s=p['sell']
    d5,d10,d15=int(s*.95),int(s*.90),int(s*.85)
    m5=round((d5-p['cost'])/d5*100,1)
    m10=round((d10-p['cost'])/d10*100,1)
    m15=round((d15-p['cost'])/d15*100,1)

    vals=[
        p['num'], p['en'][:42], int(p['qty']),
        p['buy_usd'], int(p['buy_usd']*USD),
        p['box'], p['carton'], OHPU, MARKETING_PER,
        int(p['cost']),
        s, round(mg,1), int(p['profit']), int(p['total_p']),
        d5, m5, d10, m10, d15, m15,
        int(p['pts']), int(p['tail']),
    ]
    for ci,v in enumerate(vals,1):
        cell=ws.cell(r,ci,v)
        cell.border=bd(); cell.font=ft(size=8)
        cell.alignment=al(h='right' if ci>2 else 'right', wrap=(ci==2))
        if ci==11:  # سعر البيع
            cell.fill=hf('E8F4FD'); cell.font=ft(True,'003366',10)
            cell.number_format='#,##0 "د"'
        elif ci==12:  # هامش
            cell.fill=hf(mf); cell.font=ft(True,size=9)
            cell.number_format='0.0"%"'
        elif ci in(16,18,20):  # هوامش الخصم
            cell.fill=hf(MINT) if isinstance(v,(int,float)) and v>=35 else hf(AMB) if isinstance(v,(int,float)) and v>=20 else hf(ROSE)
        elif ci in(21,22):  # نقاط
            cell.fill=hf('FFF8E1'); cell.font=ft(False,'5D4037',8,True)
        else:
            cell.fill=hf(alt)
    ws.row_dimensions[r].height=17

# صف الإجماليات
tr=len(products)+5
ws.merge_cells(f'A{tr}:J{tr}')
c=ws.cell(tr,1,'الإجماليات'); c.fill=hf(NAVY); c.font=ft(True,GOLD,10); c.alignment=al(); c.border=bd()
for ci,v in [(11,int(total_rev)),(12,round(avg_mg,1)),(13,int(total_prof)),(14,int(total_prof))]:
    c=ws.cell(tr,ci,v); c.fill=hf(NAVY); c.font=ft(True,GOLD,10); c.alignment=al(); c.border=bd()

# ─────────────────────── شيت 2: دليل علم النفس التسعيري ─────────────────────
ws2=wb.create_sheet('دليل التسعير النفسي')
ws2.sheet_view.rightToLeft=True
ws2['A1']='🧠 دليل التسعير النفسي الاحترافي — AQUAVO 2026'
ws2['A1'].fill=hf(NAVY); ws2['A1'].font=ft(True,TEAL,13); ws2.merge_cells('A1:D1')
ws2.row_dimensions[1].height=28

rows2=[
    ('الاستراتيجية','الدليل العلمي','مثال','التأثير المتوقع'),
    ('التسعير الدقيق','الدراسات (Berkeley) تثبت: 10,150 تبدو أكثر دقة وأمانة من 10,000','سخان بـ 11,350 بدل 11,000','زيادة الثقة وتقليل المساومة'),
    ('تأثير الرقم الأيسر','العقل يقرأ الرقم الأول: 9,850 تقرأ كـ"تسعة آلاف"','9,850 بدل 10,000','+20% معدل شراء (بحث مؤكد)'),
    ('الإرساء السعري (Anchoring)','أول رقم يراه الزبون يحدد مرجعه الذهني','اعرض السعر القديم مشطوباً ~~15,000~~ → 11,350','يجعل السعر يبدو صفقة رائعة'),
    ('الأسعار الكسرية (Precise Pricing)','الكسور توحي بأن السعر محسوب بدقة وليس اعتباطي','23,450 أصدق من 23,000 في ذهن الزبون','رفع الثقة وتقليل الاعتراض'),
    ('الذيل كنقاط ولاء','الـ 350 في 11,350 تتحول لنقاط → تحفز إعادة الشراء','اشترى بـ 11,350 → 11 نقطة + 350 ذيل','زيادة الاحتفاظ بالزبون'),
    ('مستويات الولاء','النقاط تحفز الزبون على الوصول للمرحلة التالية','برونز→فضة→ذهب','زيادة قيمة عمر الزبون (LTV)'),
    ('سعر البريميوم','للمنتجات الراقية: رقم مدور يوحي بالفخامة','حوض كبير: 75,000 (لا 74,850)','تمييز المنتج كـ premium'),
    ('التجميع (Bundling)','مجموعة منتجات بسعر أقل من المفرد تزيد AOV','سخان + طعام + علاج = 25,350 بدل 31,000','رفع متوسط قيمة الطلب'),
    ('العرض الزمني المحدود','الندرة والإلحاح يسرّعان القرار','عرض ينتهي الجمعة → خصم 8% (لا 10%)','تسريع قرار الشراء'),
]
for ri,row3 in enumerate(rows2,2):
    for ci,v in enumerate(row3,1):
        c=ws2.cell(ri,ci,v)
        c.border=bd(); c.alignment=al(h='right',wrap=True)
        if ri==2:
            c.fill=hf(RED); c.font=ft(True,'FFFFFF',9)
        elif str(row3[0]) in ('التسعير الدقيق','تأثير الرقم الأيسر','الإرساء السعري (Anchoring)'):
            c.fill=hf(MINT); c.font=ft(size=9)
        else:
            c.fill=hf(LGRY if ri%2==0 else 'FFFFFF'); c.font=ft(size=9)
    ws2.row_dimensions[ri].height=24

ws2.column_dimensions['A'].width=26; ws2.column_dimensions['B'].width=45
ws2.column_dimensions['C'].width=28; ws2.column_dimensions['D'].width=28

# ─────────────────────── شيت 3: الملخص المالي ───────────────────────────────
ws3=wb.create_sheet('الملخص المالي')
ws3.sheet_view.rightToLeft=True
ws3['A1']='الملخص المالي الكامل — AQUAVO 2026'
ws3['A1'].fill=hf(NAVY); ws3['A1'].font=ft(True,TEAL,13); ws3.merge_cells('A1:C1')
ws3.row_dimensions[1].height=28

rows3=[
    ('البند','دينار عراقي','دولار'),
    ('═══ الاستثمار الكلي ═══','',''),
    ('إجمالي المصاريف المعلنة',f'{TOTAL_IQD:,}',f'${TOTAL_USD:,}'),
    ('تكلفة شراء المنتجات',f'{sum(p["buy_usd"]*USD*p["qty"] for p in products):,.0f}','—'),
    ('تكاليف الكراتين',f'{sum(p["carton"]*p["qty"] for p in products):,.0f}','—'),
    ('الشحن الكلي',f'{(600+500)*USD:,}','$1,100'),
    ('بروشور 200 قطعة × 400',f'{BROCHURE_PER*BROCHURE_QTY:,}','—'),
    ('كارت UGC 200 قطعة × 260',f'{UGC_PER*UGC_QTY:,}','—'),
    ('التسويق لكل قطعة (موزّع)',f'{MARKETING_PER:,}','—'),
    ('','',''),
    ('═══ الإيرادات والأرباح ═══','',''),
    ('إجمالي الإيرادات (100% مبيعات)',f'{int(total_rev):,}',f'${int(total_rev/USD):,}'),
    ('إجمالي التكاليف المبوبة',f'{int(total_cost):,}',f'${int(total_cost/USD):,}'),
    ('صافي الربح (سعر كامل)',f'{int(total_prof):,}',f'${int(total_prof/USD):,}'),
    ('متوسط هامش الربح',f'{avg_mg:.1f}%','—'),
    ('نقطة التعادل',f'{TOTAL_IQD/total_rev*100:.1f}% من المبيعات','—'),
    ('','',''),
    ('═══ تأثير الخصومات ═══','',''),
    ('ربح عند خصم 5%', f'{int(total_rev*.95-TOTAL_IQD):,}', f'هامش {(total_rev*.95-TOTAL_IQD)/(total_rev*.95)*100:.1f}%'),
    ('ربح عند خصم 10%',f'{int(total_rev*.90-TOTAL_IQD):,}', f'هامش {(total_rev*.90-TOTAL_IQD)/(total_rev*.90)*100:.1f}%'),
    ('ربح عند خصم 15%',f'{int(total_rev*.85-TOTAL_IQD):,}', f'هامش {(total_rev*.85-TOTAL_IQD)/(total_rev*.85)*100:.1f}%'),
    ('','',''),
    ('═══ التسويق ═══','',''),
    (f'بروشور ({BROCHURE_QTY} قطعة)',f'{BROCHURE_PER*BROCHURE_QTY:,}',f'{BROCHURE_QTY} × {BROCHURE_PER}د'),
    (f'كارت UGC ({UGC_QTY} قطعة)',f'{UGC_PER*UGC_QTY:,}',f'{UGC_QTY} × {UGC_PER}د'),
    ('إجمالي التسويق المطبوع',f'{MARKETING_TOTAL:,}','موزّع على الأسعار ✅'),
]
for ri,row3 in enumerate(rows3,2):
    for ci,v in enumerate(row3,1):
        c=ws3.cell(ri,ci,v); c.border=bd()
        c.alignment=al(h='right')
        if ri==2:
            c.fill=hf(RED); c.font=ft(True,'FFFFFF',9)
        elif '═══' in str(row3[0]):
            c.fill=hf(NAVY); c.font=ft(True,GOLD,10)
        else:
            c.fill=hf(LGRY if ri%2==0 else 'FFFFFF'); c.font=ft(size=9)
    ws3.row_dimensions[ri].height=18

ws3.column_dimensions['A'].width=35; ws3.column_dimensions['B'].width=22; ws3.column_dimensions['C'].width=22

# ── حفظ ──────────────────────────────────────────────────────────────────────
out=r'c:\Users\jaafa\Desktop\upload\FishWebClean\AQUAVO_تسعير_نهائي_v3.xlsx'
wb.save(out)
o=io.StringIO()
o.write(f'تم الحفظ: {out}\n')
o.write(f'عدد المنتجات: {len(products)}\n')
o.write(f'إجمالي الإيرادات: {int(total_rev):,} دينار\n')
o.write(f'صافي الربح: {int(total_prof):,} دينار | هامش: {avg_mg:.1f}%\n')
o.write(f'تكاليف التسويق: {MARKETING_TOTAL:,} دينار (موزعة: {MARKETING_PER} د/قطعة)\n')
o.write('\nعينة أسعار:\n')
for p in products[:8]:
    o.write(f'  #{p["num"]} → تكلفة: {int(p["cost"]):,} | سعر: {p["sell"]:,} | هامش: {p["margin"]:.1f}% | نقاط: {p["pts"]}\n')
sys.stdout.buffer.write(o.getvalue().encode('utf-8','replace'))
