
import openpyxl, os, io, sys, math
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
USD     = 1520
TOTAL   = 2028 * USD          # 3,082,560 IQD
OHPU    = 2722                 # overhead per unit (shipping dist.)
CARTON  = {'S':500,'M':800,'L':1000}

# ── Psych pricing rules (Iraqi market research) ───────────────────────────────
# Left-digit effect: keep left digit LOW  → 9,900 beats 10,000
# Nine-ending preference confirmed in Kurdistan research
# Align with banknote denominations: 250/500/1000/5000/25000
# For value goods:  end in 900, 9500, 9750, 4900, etc.
# For premium goods: end in 000 or 500 (prestige signal)
PSYCH_TAILS_VALUE   = [900, 9500, 4900, 9900, 4500, 750, 250, 9750, 4750]
PSYCH_TAILS_PREMIUM = [0, 500, 5000]

CAT_MAP = {
    'heater':  ('premium', 0.42),
    'filter':  ('premium', 0.42),
    'tank':    ('premium', 0.45),
    'food':    ('value',   0.50),
    'medical': ('value',   0.55),
    'chem':    ('value',   0.52),
    'other':   ('value',   0.45),
}

def detect(en, cn):
    s = (en+cn).lower()
    if any(k in s for k in ['heat','ysh','quartz','加热']): return 'heater'
    if any(k in s for k in ['filter','pump','ylc','oil film','滤']): return 'filter'
    if any(k in s for k in ['tank','mm thick','缸','glass']): return 'tank'
    if any(k in s for k in ['feed','food','brine','worm','feast','饲料','粮']): return 'food'
    if any(k in s for k in ['spot','methylene','bactericid','白点','黄粉']): return 'medical'
    if any(k in s for k in ['probio','nitrif','stabilizer','algae','chlorine',
                              'ammonia','mineral','bacteria','test','菌','盐','除','净']): return 'chem'
    return 'other'

def psych_price(cost, margin, cat):
    """Return psychologically optimised price ≥ cost/(1-margin)."""
    min_p  = cost / (1 - margin)
    style, _ = CAT_MAP[cat]
    tails = PSYCH_TAILS_PREMIUM if style == 'premium' else PSYCH_TAILS_VALUE

    best = None
    # Try multiples of 1000 then 500 then 250
    for base_unit in [1000, 500, 250, 100]:
        candidate_base = int(math.ceil(min_p / base_unit) * base_unit)
        for tail in tails:
            # build candidate: e.g. 12000 + 900 = 12,900  OR 13000 - 100 = 12,900
            if tail < base_unit:
                c = candidate_base - base_unit + tail
            else:
                c = (candidate_base // 1000) * 1000 + tail
            if c < min_p:
                c += 1000
            if c >= min_p:
                if best is None or c < best:
                    best = c
        if best:
            break
    return int(best) if best else int(math.ceil(min_p / 100) * 100)

# ── Read source (Yee_Products_2026_UPDATED) ───────────────────────────────────
src_path = r'c:\Users\jaafa\Desktop\upload\FishWebClean\Yee_Products_2026_UPDATED.xlsx'
wb_src   = openpyxl.load_workbook(src_path, data_only=True)
src_rows = list(wb_src.active.iter_rows(values_only=True))

# ── Read packing list 1.9 普货 ─────────────────────────────────────────────────
pl_reg_path = r'C:\Users\jaafa\Downloads\ros\伊拉克-Jaafar-1.9 - packing list（普货）.xlsx'
wb_reg = openpyxl.load_workbook(pl_reg_path, data_only=True)
reg_codes = set()
for row in wb_reg.active.iter_rows(values_only=True):
    if row[2] and str(row[2]).strip() not in ('Business\nCode',''):
        reg_codes.add(str(row[2]).strip())
        reg_codes.add(str(row[3]).strip() if row[3] else '')

# ── Read packing list 1.9 敏货 ─────────────────────────────────────────────────
pl_sen_path = r'C:\Users\jaafa\Downloads\ros\伊拉克-Jaafar-1.9 - packing list（敏货）.xlsx'
wb_sen = openpyxl.load_workbook(pl_sen_path, data_only=True)
sen_codes = set()
for row in wb_sen.active.iter_rows(values_only=True):
    if row[2] and str(row[2]).strip() not in ('Business\nCode',''):
        sen_codes.add(str(row[2]).strip())
        sen_codes.add(str(row[3]).strip() if row[3] else '')

all_pl_codes = reg_codes | sen_codes

# ── Build product list from source ────────────────────────────────────────────
products = []
for row in src_rows[1:]:
    if not row[0]: continue
    try:
        qty     = float(str(row[3]).replace(',','') or 0)
        buy_usd = float(str(row[4]).replace(',','') or 0)
        box     = str(row[5]) if row[5] else 'S'
        carton  = CARTON.get(box, 500)
        en      = str(row[2]) if row[2] else ''
        cn      = str(row[1]) if row[1] else ''
        cat     = detect(en, cn)
        _, min_m = CAT_MAP[cat]
        cost    = buy_usd * USD + carton + OHPU
        sell    = psych_price(cost, min_m, cat)
        margin  = (sell - cost) / sell * 100
        profit  = sell - cost
        pts     = sell // 1000
        tail    = sell % 1000
        products.append({
            'num':row[0],'en':en,'cn':cn,'qty':qty,'buy_usd':buy_usd,
            'box':box,'carton':carton,'cost':cost,'sell':sell,
            'margin':margin,'profit':profit,'total_p':profit*qty,
            'pts':pts,'tail':tail,'cat':cat,
        })
    except Exception as e:
        pass

total_rev  = sum(p['sell']*p['qty'] for p in products)
total_cost = sum(p['cost']*p['qty'] for p in products)
total_prof = sum(p['total_p'] for p in products)
avg_margin = total_prof / total_rev * 100

# ── Styles ────────────────────────────────────────────────────────────────────
def hf(c): return PatternFill('solid', fgColor=c)
def ft(bold=False, color='000000', size=9, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name='Calibri')
def al(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
T = Side(style='thin', color='CCCCCC')
def bd(): return Border(top=T, bottom=T, left=T, right=T)

NAVY='1A1A2E'; TEAL='A8DADC'; RED='E63946'; GOLD='FFD700'
MINT='D8F3DC'; AMB='FFF3CD'; ROSE='FFD6D6'; LGREY='F8F9FA'

# ═══════════════════════════════════════════════════════════════════════════
# WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ── SHEET 1: Pricing (with variable margin input) ──────────────────────────
ws = wb.active
ws.title = 'Pricing'
ws.freeze_panes = 'A5'

# Row 1: Title
ws.merge_cells('A1:S1')
c=ws['A1']; c.value='AQUAVO — YEE Pricing 2026  |  1 USD = 1,520 IQD  |  Total Cost: $2,028'
c.fill=hf(NAVY); c.font=ft(True,TEAL,12); c.alignment=al()
ws.row_dimensions[1].height=26

# Row 2: Margin selector instructions
ws.merge_cells('A2:S2')
c=ws['A2']; c.value='👆 اختر نسبة ربحك في الخلية U2 (مثال: 0.45 = 45%) — الأسعار تتحدث تلقائياً  |  Psych Pricing: Iraqi market optimised (9,900 / 14,500 / 7,900...)'
c.fill=hf('16213E'); c.font=ft(False,TEAL,9,True); c.alignment=al()
ws.row_dimensions[2].height=16

# Row 3: User input cell
ws['U3']='نسبة الربح المختارة'
ws['U3'].font=ft(True,RED,10); ws['U3'].alignment=al()
ws['V3']=0.45
ws['V3'].font=ft(True,GOLD,11); ws['V3'].alignment=al()
ws['V3'].number_format='0%'
ws['V3'].fill=hf('FFF0D0')
ws.column_dimensions['U'].width=22; ws.column_dimensions['V'].width=10

# Row 4: Headers
HEADERS=[
    ('#',4),('Product EN',38),('اسم المنتج',20),('Qty',5),
    ('Buy $',7),('Buy IQD',9),('Box',4),('Carton',8),('Ship/u',8),
    ('Cost/u',10),('Sell Price\n(Psych)',11),('Margin%',8),
    ('Profit/u',9),('Tot Profit',10),
    ('5% disc\nprice',9),('5%\nmargin',7),
    ('10% disc\nprice',9),('10%\nmargin',7),
    ('15% disc\nprice',9),('15%\nmargin',7),
    ('Loyalty\nPts',8),('Tail\nIQD',7),
]
for ci,(h,w) in enumerate(HEADERS,1):
    c=ws.cell(3,ci,h)
    c.fill=hf(RED); c.font=ft(True,'FFFFFF',8); c.alignment=al(wrap=True); c.border=bd()
    ws.column_dimensions[get_column_letter(ci)].width=w
ws.row_dimensions[3].height=30

# Data rows
for i,p in enumerate(products):
    r=i+4; alt=LGREY if i%2==0 else 'FFFFFF'
    mg=p['margin']
    mg_fill = MINT if mg>=48 else AMB if mg>=35 else ROSE

    s=p['sell']
    vals=[
        p['num'], p['en'][:40], p['cn'][:20],
        int(p['qty']), p['buy_usd'], int(p['buy_usd']*USD),
        p['box'], p['carton'], OHPU,
        int(p['cost']), s,
        round(mg,1), int(p['profit']), int(p['total_p']),
        int(s*0.95), round((s*0.95-p['cost'])/(s*0.95)*100,1),
        int(s*0.90), round((s*0.90-p['cost'])/(s*0.90)*100,1),
        int(s*0.85), round((s*0.85-p['cost'])/(s*0.85)*100,1),
        int(p['pts']), int(p['tail']),
    ]
    for ci,v in enumerate(vals,1):
        cell=ws.cell(r,ci,v)
        cell.border=bd(); cell.font=ft(size=8)
        cell.alignment=al(h='right' if ci>3 else 'left', wrap=(ci<=3))
        if ci==11:
            cell.fill=hf('E8F4FD'); cell.font=ft(True,'003366',9)
            cell.number_format='#,##0 "IQD"'
        elif ci==12:
            cell.fill=hf(mg_fill); cell.font=ft(True,size=9)
        elif ci in(16,18,20):
            v2=vals[ci-1]
            cell.fill=hf(MINT) if isinstance(v2,float) and v2>=35 else hf(AMB) if isinstance(v2,float) and v2>=20 else hf(ROSE)
        elif ci in(21,22):
            cell.fill=hf('FFF8E1'); cell.font=ft(False,'5D4037',8,True)
        else:
            cell.fill=hf(alt)
    ws.row_dimensions[r].height=16

# Totals row
tr=len(products)+4
ws.merge_cells(f'A{tr}:J{tr}')
c=ws.cell(tr,1,'TOTALS'); c.fill=hf(NAVY); c.font=ft(True,GOLD,10); c.alignment=al()
for ci,v in [(11,int(total_rev)),(12,round(avg_margin,1)),(13,int(total_rev-total_cost)),(14,int(total_prof))]:
    c=ws.cell(tr,ci,v); c.fill=hf(NAVY); c.font=ft(True,GOLD,10); c.alignment=al(); c.border=bd()

# ── SHEET 2: Match Check ──────────────────────────────────────────────────────
ws2=wb.create_sheet('Match Check 1.9')
ws2['A1']='PACKING LIST 1.9 vs UPDATED FILE — Verification'
ws2['A1'].fill=hf(NAVY); ws2['A1'].font=ft(True,TEAL,12)
ws2.merge_cells('A1:F1'); ws2.row_dimensions[1].height=24

heads=['#','Product EN','Chinese Name','In 普货 1.9?','In 敏货 1.9?','Status']
widths=[4,38,22,14,14,16]
for ci,(h,w) in enumerate(zip(heads,widths),1):
    c=ws2.cell(2,ci,h); c.fill=hf(RED); c.font=ft(True,'FFFFFF',9); c.alignment=al(); c.border=bd()
    ws2.column_dimensions[get_column_letter(ci)].width=w
ws2.row_dimensions[2].height=20

# Build check: match by Chinese name or code snippet
reg_names={str(r[3]).strip() for r in wb_reg.active.iter_rows(values_only=True) if r[3]}
reg_en={str(r[4]).strip()[:20] for r in wb_reg.active.iter_rows(values_only=True) if r[4]}
sen_names={str(r[3]).strip() for r in wb_sen.active.iter_rows(values_only=True) if r[3]}
sen_en={str(r[4]).strip()[:20] for r in wb_sen.active.iter_rows(values_only=True) if r[4]}

missing=[]
for i,p in enumerate(products):
    r=i+3
    in_reg = p['cn'][:20] in reg_names or p['en'][:20] in reg_en
    in_sen = p['cn'][:20] in sen_names or p['en'][:20] in sen_en
    if not in_reg and not in_sen:
        missing.append(p['en'][:40])
    status = 'OK' if (in_reg or in_sen) else 'NOT FOUND'
    s_fill = MINT if status=='OK' else ROSE
    vals2=[p['num'],p['en'][:40],p['cn'][:20],
           'YES' if in_reg else '—','YES' if in_sen else '—',status]
    for ci,v in enumerate(vals2,1):
        c=ws2.cell(r,ci,v); c.border=bd(); c.font=ft(size=8)
        c.alignment=al(h='left' if ci in(2,3) else 'center')
        c.fill=hf(s_fill if ci==6 else (LGREY if i%2==0 else 'FFFFFF'))
    ws2.row_dimensions[r].height=16

# ── SHEET 3: Summary ──────────────────────────────────────────────────────────
ws3=wb.create_sheet('Summary')
ws3['A1']='AQUAVO Financial Summary — 1 USD = 1,520 IQD'
ws3['A1'].fill=hf(NAVY); ws3['A1'].font=ft(True,TEAL,12)
ws3.merge_cells('A1:D1'); ws3.row_dimensions[1].height=24

rows3=[
    ('INVESTMENT','',''),
    ('Total Expenses','3,082,560 IQD','$2,028'),
    ('Product Cost','1,335,168 IQD','$878'),
    ('Cartons','318,100 IQD','—'),
    ('Shipping','1,716,000 IQD','$1,100'),
    ('','',''),
    ('REVENUE (full price)','',''),
    ('Total Revenue',f'{int(total_rev):,} IQD',f'${int(total_rev/USD):,}'),
    ('Total Profit',f'{int(total_prof):,} IQD',f'${int(total_prof/USD):,}'),
    ('Avg Margin',f'{avg_margin:.1f}%',''),
    ('Break-even',f'{TOTAL/total_rev*100:.1f}% of items',''),
    ('','',''),
    ('DISCOUNT SCENARIOS','',''),
    ('5% disc profit',f'{int(total_rev*0.95-TOTAL):,} IQD',f'margin {(total_rev*0.95-TOTAL)/(total_rev*0.95)*100:.1f}%'),
    ('10% disc profit',f'{int(total_rev*0.90-TOTAL):,} IQD',f'margin {(total_rev*0.90-TOTAL)/(total_rev*0.90)*100:.1f}%'),
    ('15% disc profit',f'{int(total_rev*0.85-TOTAL):,} IQD',f'margin {(total_rev*0.85-TOTAL)/(total_rev*0.85)*100:.1f}%'),
    ('','',''),
    ('MISSING FROM PACKING LIST 1.9','',''),
]
for mi,m in enumerate(missing):
    rows3.append((f'  {m}','NOT IN 1.9',''))

for ri,row3 in enumerate(rows3,2):
    for ci,v in enumerate(row3,1):
        c=ws3.cell(ri,ci,v)
        c.font=ft(True,RED,10) if str(v) in('INVESTMENT','REVENUE (full price)','DISCOUNT SCENARIOS','MISSING FROM PACKING LIST 1.9') else ft(size=9)
        c.alignment=al(h='left'); c.border=bd()
        c.fill=hf('FFF0F0') if str(row3[0]).isupper() and row3[0] else hf(LGREY if ri%2==0 else 'FFFFFF')
    ws3.row_dimensions[ri].height=16

ws3.column_dimensions['A'].width=38
ws3.column_dimensions['B'].width=20
ws3.column_dimensions['C'].width=18

# ── SHEET 4: Psych Pricing Guide ─────────────────────────────────────────────
ws4=wb.create_sheet('Psych Pricing Guide')
ws4['A1']='دليل التسعير النفسي — AQUAVO Iraq 2026'
ws4['A1'].fill=hf(NAVY); ws4['A1'].font=ft(True,TEAL,12)
ws4.merge_cells('A1:C1')

guide4=[
    ('الاستراتيجية','السبب العلمي','مثال عملي'),
    ('9,900 بدل 10,000','Left-digit effect: العقل يقرأ 9 لا 10 → يحس أرخص','سخان 9,900 IQD'),
    ('14,500 بدل 15,000','Charm pricing: انتهاء بـ 500 يوحي بدقة وعدالة','علاج 14,500 IQD'),
    ('7,900 بدل 8,000','Nine-ending proven in Kurdistan research +20% sales','طعام 7,900 IQD'),
    ('24,500 (premium)','Even hundreds = ثقة وجودة عند المنتجات الغالية','حوض 24,500 IQD'),
    ('السعر الأصلي مشطوب','Price anchoring: ~~15,000~~ ← 9,900 يحفز الشراء','عرض لافت'),
    ('نهاية بـ 750','الـ 750 تصير نقاط ولاء تحفز إعادة الشراء','11,750 → 11 نقطة+750'),
    ('نهاية بـ 900','الـ 100 الباقي يشعر بـ"باقي يرجع" → ولاء','9,900 → 9 نقاط'),
    ('سعر بدون تقريب','الأرقام غير المدورة تبدو أكثر دقة ومصداقية','7,450 أدق من 7,500'),
]
for ri,row4 in enumerate(guide4,2):
    for ci,v in enumerate(row4,1):
        c=ws4.cell(ri,ci,v)
        c.font=ft(True,RED,10) if ri==2 else ft(size=9)
        c.fill=hf(NAVY) if ri==2 else hf(LGREY if ri%2==0 else 'FFFFFF')
        if ri==2: c.font=ft(True,'FFFFFF',9)
        c.border=bd(); c.alignment=al(h='left',wrap=True)
    ws4.row_dimensions[ri].height=22
ws4.column_dimensions['A'].width=22
ws4.column_dimensions['B'].width=38
ws4.column_dimensions['C'].width=28

# ── Save ─────────────────────────────────────────────────────────────────────
out_path=r'c:\Users\jaafa\Desktop\upload\FishWebClean\AQUAVO_Pricing_v2_FINAL.xlsx'
wb.save(out_path)
out=io.StringIO()
out.write(f'SAVED: {out_path}\n')
out.write(f'Products: {len(products)}\n')
out.write(f'Revenue: {int(total_rev):,} IQD\n')
out.write(f'Profit: {int(total_prof):,} IQD | Margin: {avg_margin:.1f}%\n')
out.write(f'Missing from 1.9 packing lists: {len(missing)}\n')
for m in missing: out.write(f'  - {m}\n')
sys.stdout.buffer.write(out.getvalue().encode('utf-8','replace'))
